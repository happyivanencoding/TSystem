from __future__ import annotations

from functools import reduce
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd


TP_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FS_SECTOR_WORKBOOK_DIR = TP_ROOT / "13_sector_score_model"

DATE_COLUMN = "Date"
MARKET_COLUMN = "FS_MARKET_FS_SECTOR"
SECTOR_NAME_COLUMN = "FS_SECTOR_NAME_FS_SECTOR"
SECTOR_CODE_COLUMN = " Benchmark ICB Supersector "
US_WEIGHT_COLUMN = "Weight in SP500"
EU_WEIGHT_COLUMN = "Weight in STOXX EUROPE 600"

KEY_COLUMNS = [DATE_COLUMN, MARKET_COLUMN, SECTOR_NAME_COLUMN]

US_EXCEL_SECTORS = [
    "Materials",
    "ConsStaples",
    "Retail",
    "Fin",
    "HealthCare",
    "Indus",
    "Oil",
    "Tech",
    "Telco",
    "Utili",
    "Travel & leisure",
    "Media",
]

EU_EXCEL_SECTORS = [
    "Materials",
    "ConsStaples",
    "Pers. Goods",
    "Fin",
    "HealthCare",
    "Indus",
    "Oil",
    "Tech",
    "Telco",
    "Utili",
    "Travel & leisure",
    "Media",
    "Auto",
]

US_ICB19_TO_EXCEL_SECTOR = {
    2: "Fin",
    3: "Materials",
    4: "Materials",
    5: "Indus",
    6: "Fin",
    7: "ConsStaples",
    8: "HealthCare",
    9: "Indus",
    10: "Fin",
    11: "Media",
    12: "Oil",
    13: "ConsStaples",
    14: "Fin",
    15: "Retail",
    16: "Tech",
    17: "Telco",
    18: "Travel & leisure",
    19: "Utili",
}

EU_ICB19_TO_EXCEL_SECTOR = {
    1: "Auto",
    2: "Fin",
    3: "Materials",
    4: "Materials",
    5: "Indus",
    6: "Fin",
    7: "ConsStaples",
    8: "HealthCare",
    9: "Indus",
    10: "Fin",
    11: "Media",
    12: "Oil",
    13: "Pers. Goods",
    14: "Fin",
    16: "Tech",
    17: "Telco",
    18: "Travel & leisure",
    19: "Utili",
}

MARKET_CONFIG = {
    "US": {
        "workbook": "Score_Sectoriel_US.xlsm",
        "weight_column": US_WEIGHT_COLUMN,
        "excel_sectors": US_EXCEL_SECTORS,
        "sector_map": US_ICB19_TO_EXCEL_SECTOR,
    },
    "EU": {
        "workbook": "Score_Sectoriel_EU.xlsm",
        "weight_column": EU_WEIGHT_COLUMN,
        "excel_sectors": EU_EXCEL_SECTORS,
        "sector_map": EU_ICB19_TO_EXCEL_SECTOR,
    },
}

FMA_SHEETS = {
    "Leverage_FMA": "LEVERAGE",
    "Margin_FMA": "MARGIN",
    "Valuation_FMA_hist": "VALUE",
    "MOM_FMA": "MOMENTUM",
    "Growth_FMA": "GROWTH",
    "Vol_FMA": "LOW_VOL",
}

FS_SECTOR_DATA_COLUMNS = [
    MARKET_COLUMN,
    SECTOR_NAME_COLUMN,
    "LEVERAGE_RANK_FS_SECTOR",
    "LEVERAGE_SCORE_FS_SECTOR",
    "MARGIN_RANK_FS_SECTOR",
    "MARGIN_SCORE_FS_SECTOR",
    "VALUE_RANK_FS_SECTOR",
    "VALUE_SCORE_FS_SECTOR",
    "MOMENTUM_RANK_FS_SECTOR",
    "MOMENTUM_SCORE_FS_SECTOR",
    "GROWTH_RANK_FS_SECTOR",
    "GROWTH_SCORE_FS_SECTOR",
    "LOW_VOL_RANK_FS_SECTOR",
    "LOW_VOL_SCORE_FS_SECTOR",
    "FIVE_FACTOR_RANK_FS_SECTOR",
    "FIVE_FACTOR_SCORE_FS_SECTOR",
    "RECO_TOP_FLAG_FS_SECTOR",
    "RECO_WORST_FLAG_FS_SECTOR",
    "RECO_SCORE_FS_SECTOR",
    "BENCH_WEIGHT_FS_SECTOR",
    "MACRO_SIGNAL_AVG_FS_SECTOR",
    "MACRO_CYCLE_CODE_FS_SECTOR",
    "RATE_SIGNAL_FS_SECTOR",
]


def _ensure_isin_column(df: pd.DataFrame) -> pd.DataFrame:
    if "ISIN" in df.columns:
        return df.copy()
    if df.index.name == "ISIN":
        return df.reset_index()
    raise ValueError("DataFrame must contain ISIN as a column or index")


def _month_end(value: Any) -> pd.Timestamp | pd.NaT:
    timestamp = pd.to_datetime(value, errors="coerce")
    if pd.isna(timestamp):
        return pd.NaT
    return pd.Timestamp(timestamp) + pd.offsets.MonthEnd(0)


def _to_float(value: Any) -> float:
    if value is None:
        return float("nan")
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "#N/A", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?"}:
            return float("nan")
        value = stripped
    parsed = pd.to_numeric(value, errors="coerce")
    return float(parsed) if pd.notna(parsed) else float("nan")


def _clean_text(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", "#N/A", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?"}:
            return None
        return stripped
    return value


def _normalize_label(value: Any) -> str:
    return str(value).strip().upper() if value is not None else ""


def _find_in_row(ws: Any, row: int, labels: set[str]) -> int:
    normalized_labels = {_normalize_label(label) for label in labels}
    for column in range(1, ws.max_column + 1):
        if _normalize_label(ws.cell(row, column).value) in normalized_labels:
            return column
    raise ValueError(f"Could not find labels {sorted(labels)} in {ws.title} row {row}")


def _extract_sector_block(
    ws: Any,
    market: str,
    sectors: list[str],
    feature_column: str,
    value_start_column: int,
    date_column: int,
    first_data_row: int,
    include_nan: bool = False,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for row in range(first_data_row, ws.max_row + 1):
        date = _month_end(ws.cell(row, date_column).value)
        if pd.isna(date):
            continue
        for offset, sector in enumerate(sectors):
            value = _to_float(ws.cell(row, value_start_column + offset).value)
            if not include_nan and np.isnan(value):
                continue
            records.append(
                {
                    DATE_COLUMN: date,
                    MARKET_COLUMN: market,
                    SECTOR_NAME_COLUMN: sector,
                    feature_column: value,
                }
            )
    return pd.DataFrame.from_records(records)


def _merge_feature_frames(frames: list[pd.DataFrame]) -> pd.DataFrame:
    frames = [frame for frame in frames if not frame.empty]
    if not frames:
        return pd.DataFrame(columns=KEY_COLUMNS)
    merged = reduce(lambda left, right: left.merge(right, on=KEY_COLUMNS, how="outer"), frames)
    return merged.drop_duplicates(subset=KEY_COLUMNS, keep="last")


def _extract_fma_features(wb: Any, market: str, sectors: list[str]) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    for sheet_name, prefix in FMA_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        frames.append(
            _extract_sector_block(
                ws,
                market,
                sectors,
                f"{prefix}_RANK_FS_SECTOR",
                value_start_column=2,
                date_column=1,
                first_data_row=8,
            )
        )
        score_start_column = _find_in_row(
            ws,
            3,
            {"5Y AVERAGE PERCENTILE", "AVERAGE 1-10 SCORE"},
        )
        frames.append(
            _extract_sector_block(
                ws,
                market,
                sectors,
                f"{prefix}_SCORE_FS_SECTOR",
                value_start_column=score_start_column,
                date_column=score_start_column - 1,
                first_data_row=8,
            )
        )
    return frames


def _extract_five_factor_features(wb: Any, market: str, sectors: list[str]) -> list[pd.DataFrame]:
    if "5F_FMA" not in wb.sheetnames:
        return []
    ws = wb["5F_FMA"]
    weighted_score_column = _find_in_row(ws, 3, {"Weighted Score"})
    return [
        _extract_sector_block(
            ws,
            market,
            sectors,
            "FIVE_FACTOR_RANK_FS_SECTOR",
            value_start_column=2,
            date_column=1,
            first_data_row=8,
        ),
        _extract_sector_block(
            ws,
            market,
            sectors,
            "FIVE_FACTOR_SCORE_FS_SECTOR",
            value_start_column=weighted_score_column,
            date_column=1,
            first_data_row=8,
        ),
    ]


def _extract_reco_features(wb: Any, market: str, sectors: list[str]) -> pd.DataFrame:
    if "Reco_histo" not in wb.sheetnames:
        return pd.DataFrame(columns=KEY_COLUMNS)
    ws = wb["Reco_histo"]
    top_start = _find_in_row(ws, 1, {"Top"})
    worst_start = _find_in_row(ws, 1, {"Worst"})
    top = _extract_sector_block(
        ws,
        market,
        sectors,
        "RECO_TOP_FLAG_FS_SECTOR",
        value_start_column=top_start,
        date_column=1,
        first_data_row=3,
        include_nan=True,
    )
    worst = _extract_sector_block(
        ws,
        market,
        sectors,
        "RECO_WORST_FLAG_FS_SECTOR",
        value_start_column=worst_start,
        date_column=1,
        first_data_row=3,
        include_nan=True,
    )
    reco = _merge_feature_frames([top, worst])
    if not reco.empty:
        reco["RECO_SCORE_FS_SECTOR"] = (
            reco["RECO_TOP_FLAG_FS_SECTOR"].fillna(0) - reco["RECO_WORST_FLAG_FS_SECTOR"].fillna(0)
        )
    return reco


def _extract_bench_features(wb: Any, market: str, sectors: list[str]) -> pd.DataFrame:
    if "Bench" not in wb.sheetnames:
        return pd.DataFrame(columns=KEY_COLUMNS)
    ws = wb["Bench"]
    return _extract_sector_block(
        ws,
        market,
        sectors,
        "BENCH_WEIGHT_FS_SECTOR",
        value_start_column=2,
        date_column=1,
        first_data_row=2,
    )


def _extract_macro_features(wb: Any, market: str) -> pd.DataFrame:
    if "Cycle macro" not in wb.sheetnames:
        return pd.DataFrame(columns=[DATE_COLUMN, MARKET_COLUMN])
    ws = wb["Cycle macro"]
    macro_column = _find_in_row(ws, 2, {"Moyenne"})
    cycle_column = _find_in_row(ws, 2, {"Cycle"})
    rate_signal_column = _find_in_row(ws, 3, {"Singal taux", "Signal taux"})
    records: list[dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        date = _month_end(ws.cell(row, 1).value)
        if pd.isna(date):
            continue
        records.append(
            {
                DATE_COLUMN: date,
                MARKET_COLUMN: market,
                "MACRO_SIGNAL_AVG_FS_SECTOR": _to_float(ws.cell(row, macro_column).value),
                "MACRO_CYCLE_CODE_FS_SECTOR": _clean_text(ws.cell(row, cycle_column).value),
                "RATE_SIGNAL_FS_SECTOR": _clean_text(ws.cell(row, rate_signal_column).value),
            }
        )
    return pd.DataFrame.from_records(records).drop_duplicates(
        subset=[DATE_COLUMN, MARKET_COLUMN],
        keep="last",
    )


def extract_fs_sector_history(
    workbook_dir: Path = DEFAULT_FS_SECTOR_WORKBOOK_DIR,
) -> pd.DataFrame:
    workbook_dir = Path(workbook_dir)
    market_frames: list[pd.DataFrame] = []
    for market, config in MARKET_CONFIG.items():
        workbook_path = workbook_dir / str(config["workbook"])
        if not workbook_path.exists():
            raise FileNotFoundError(f"Missing FactSet sector workbook: {workbook_path}")
        wb = openpyxl.load_workbook(
            workbook_path,
            read_only=False,
            data_only=True,
            keep_vba=False,
        )
        sectors = list(config["excel_sectors"])
        feature_frames = [
            *_extract_fma_features(wb, market, sectors),
            *_extract_five_factor_features(wb, market, sectors),
            _extract_reco_features(wb, market, sectors),
            _extract_bench_features(wb, market, sectors),
        ]
        market_features = _merge_feature_frames(feature_frames)
        macro = _extract_macro_features(wb, market)
        if not macro.empty and not market_features.empty:
            market_features = market_features.merge(macro, on=[DATE_COLUMN, MARKET_COLUMN], how="left")
        market_frames.append(market_features)
    result = pd.concat(market_frames, ignore_index=True)
    result[DATE_COLUMN] = pd.to_datetime(result[DATE_COLUMN])
    return result.sort_values(KEY_COLUMNS).reset_index(drop=True)


def _assign_fs_sector_keys(screen: pd.DataFrame) -> pd.DataFrame:
    keyed = screen.copy()
    keyed[DATE_COLUMN] = pd.to_datetime(keyed[DATE_COLUMN], errors="coerce") + pd.offsets.MonthEnd(0)
    sector_code = pd.to_numeric(keyed[SECTOR_CODE_COLUMN], errors="coerce")
    keyed[MARKET_COLUMN] = pd.NA
    keyed[SECTOR_NAME_COLUMN] = pd.NA

    us_mask = pd.Series(False, index=keyed.index)
    if US_WEIGHT_COLUMN in keyed.columns:
        us_mask = pd.to_numeric(keyed[US_WEIGHT_COLUMN], errors="coerce").fillna(0).gt(0)
        keyed.loc[us_mask, MARKET_COLUMN] = "US"
        keyed.loc[us_mask, SECTOR_NAME_COLUMN] = sector_code.loc[us_mask].map(US_ICB19_TO_EXCEL_SECTOR)

    if EU_WEIGHT_COLUMN in keyed.columns:
        eu_mask = pd.to_numeric(keyed[EU_WEIGHT_COLUMN], errors="coerce").fillna(0).gt(0)
        eu_mask = eu_mask & keyed[MARKET_COLUMN].isna()
        keyed.loc[eu_mask, MARKET_COLUMN] = "EU"
        keyed.loc[eu_mask, SECTOR_NAME_COLUMN] = sector_code.loc[eu_mask].map(EU_ICB19_TO_EXCEL_SECTOR)

    return keyed


def apply_fs_sector_history_to_frame(
    screen_df: pd.DataFrame,
    workbook_dir: Path = DEFAULT_FS_SECTOR_WORKBOOK_DIR,
    processor: Any | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    screen = _ensure_isin_column(screen_df).reset_index(drop=True)
    rows_before = int(len(screen))
    columns_before = set(screen.columns)
    fs_history = extract_fs_sector_history(Path(workbook_dir))
    feature_columns = [column for column in fs_history.columns if column not in KEY_COLUMNS]

    for column in FS_SECTOR_DATA_COLUMNS:
        if column not in screen.columns:
            screen[column] = pd.NA
        else:
            screen[column] = pd.NA

    screen = _assign_fs_sector_keys(screen)
    screen["__row_id"] = np.arange(len(screen))
    eligible = screen[MARKET_COLUMN].notna() & screen[SECTOR_NAME_COLUMN].notna()
    merge_keys = ["__row_id", *KEY_COLUMNS]
    payload = screen.loc[eligible, merge_keys].merge(fs_history, on=KEY_COLUMNS, how="left")
    matched_mask = payload[feature_columns].notna().any(axis=1) if feature_columns else pd.Series(False)

    for column in feature_columns:
        row_ids = payload["__row_id"].to_numpy()
        screen.loc[row_ids, column] = payload[column].to_numpy()

    latest_date = pd.to_datetime(screen[DATE_COLUMN]).max()
    latest = screen.loc[screen[DATE_COLUMN].eq(latest_date)].copy()
    latest_eligible = latest[MARKET_COLUMN].notna() & latest[SECTOR_NAME_COLUMN].notna()
    latest_features = latest.loc[latest_eligible, feature_columns]
    latest_matched = (
        int(latest_features.notna().any(axis=1).sum()) if feature_columns and len(latest_features) else 0
    )

    output = screen.drop(columns=["__row_id"]).set_index("ISIN")
    if processor is not None:
        processor.validate_unique_keys(output)

    if int(len(output)) != rows_before:
        raise ValueError(f"FS sector merge changed row count: before={rows_before}, after={len(output)}")

    result = {
        "workbook_dir": str(Path(workbook_dir)),
        "source_workbooks": [str(Path(workbook_dir) / str(config["workbook"])) for config in MARKET_CONFIG.values()],
        "history_rows": int(len(fs_history)),
        "history_date_min": fs_history[DATE_COLUMN].min() if len(fs_history) else None,
        "history_date_max": fs_history[DATE_COLUMN].max() if len(fs_history) else None,
        "eligible_rows": int(eligible.sum()),
        "matched_rows": int(matched_mask.sum()),
        "latest_date": latest_date,
        "latest_eligible_rows": int(latest_eligible.sum()),
        "latest_matched_rows": latest_matched,
        "new_columns": sorted(set(output.reset_index().columns) - columns_before),
        "fs_sector_columns": FS_SECTOR_DATA_COLUMNS,
        "output_rows": int(len(output)),
        "output_columns": int(len(output.columns)),
        "idempotency_mode": "overwrite FS_SECTOR columns from current workbooks; keep row count unchanged",
    }
    return output, result
