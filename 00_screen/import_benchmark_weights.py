"""Import historical benchmark weights from Bloomberg point-in-time workbooks.

The source workbooks identify securities with Bloomberg equity tickers while the
canonical screen is keyed by ``(ISIN, Date)``.  A Bloomberg-enabled mapping
workbook is therefore required before this module is allowed to write.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook

TP_ROOT = Path(__file__).resolve().parents[1]
if str(TP_ROOT) not in sys.path:
    sys.path.insert(0, str(TP_ROOT))
import sitecustomize  # noqa: F401,E402

from tp_core.data_contract import validate_screen_contract  # noqa: E402


SCREEN_PATH = TP_ROOT / "00_screen" / "screen_aggregate.parquet"
LAST_SCREEN_PATH = TP_ROOT / "00_screen" / "last_screen.parquet"
SCREEN_5Y_PATH = TP_ROOT / "00_screen" / "screen_aggregate_5Y.parquet"
INCOMING_ROOT = TP_ROOT / "00_screen" / "production_inputs" / "incoming"
QA_ROOT = TP_ROOT / "00_screen" / "qa"
MANIFEST_ROOT = TP_ROOT / "10_pipeline_runs" / "manifests" / "benchmark_weight_import"

BENCHMARK_COLUMNS = {
    "SP400": "Weight in SP400",
    "SP500": "Weight in SP500",
    "stoxx600": "Weight in STOXX EUROPE 600",
}
ISIN_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")
DATE_PATTERN = re.compile(r"Rebalance Period:\s*(.+)", re.IGNORECASE)


@dataclass(frozen=True)
class SourceRow:
    security: str
    weight: float
    short_name: str


@dataclass(frozen=True)
class Snapshot:
    benchmark: str
    source_path: Path
    source_date: pd.Timestamp
    rows: tuple[SourceRow, ...]


def _ensure_isin_column(frame: pd.DataFrame) -> pd.DataFrame:
    if "ISIN" in frame.columns:
        return frame.copy()
    if frame.index.name == "ISIN":
        return frame.reset_index()
    raise ValueError("screen 缺少 ISIN 列或 ISIN index")


def _clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_snapshot(path: Path, benchmark: str) -> Snapshot | None:
    """Read one Bloomberg point-in-time workbook; ignore count-history grids."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        title = _clean_text(sheet["A2"].value)
        if "Point In Time Results" not in title:
            return None

        match = DATE_PATTERN.search(_clean_text(sheet["D2"].value))
        if not match:
            raise ValueError(f"无法从 {path} 读取 Rebalance Period")
        source_date = pd.Timestamp(pd.to_datetime(match.group(1), errors="raise")).normalize()

        rows: list[SourceRow] = []
        for values in sheet.iter_rows(min_row=4, values_only=True):
            ticker = _clean_text(values[0] if len(values) > 0 else None)
            exchange = _clean_text(values[1] if len(values) > 1 else None)
            short_name = _clean_text(values[2] if len(values) > 2 else None)
            raw_weight = values[4] if len(values) > 4 else None
            if not ticker or not exchange:
                continue
            try:
                weight = float(raw_weight)
            except (TypeError, ValueError):
                continue
            if weight <= 0:
                continue
            rows.append(
                SourceRow(
                    security=f"{ticker} {exchange} Equity",
                    weight=weight,
                    short_name=short_name,
                )
            )

        if not rows:
            raise ValueError(f"{path} 没有可用正权重行")
        return Snapshot(benchmark, path, source_date, tuple(rows))
    finally:
        workbook.close()


def collect_snapshots(incoming_root: Path = INCOMING_ROOT) -> tuple[list[Snapshot], list[str]]:
    snapshots: list[Snapshot] = []
    skipped: list[str] = []
    for benchmark, _ in BENCHMARK_COLUMNS.items():
        directory = incoming_root / benchmark
        if not directory.is_dir():
            raise FileNotFoundError(f"缺少 benchmark 输入目录: {directory}")
        for path in sorted(directory.glob("*.xlsx")):
            snapshot = parse_snapshot(path, benchmark)
            if snapshot is None:
                skipped.append(str(path))
            else:
                snapshots.append(snapshot)

    duplicate_keys = pd.Series([(item.benchmark, item.source_date) for item in snapshots]).duplicated(keep=False)
    if duplicate_keys.any():
        duplicates = pd.Series([(item.benchmark, item.source_date) for item in snapshots])[duplicate_keys].tolist()
        raise ValueError(f"存在重复 benchmark/source_date 截面: {duplicates[:10]}")
    return snapshots, skipped


def map_source_date(source_date: pd.Timestamp, screen_dates: Iterable[pd.Timestamp]) -> pd.Timestamp:
    """Map month-start files backward; map month-end-like files to nearest month-end."""

    source_date = pd.Timestamp(source_date).normalize()
    dates = pd.DatetimeIndex(pd.to_datetime(list(screen_dates))).dropna().normalize().unique().sort_values()
    if dates.empty:
        raise ValueError("screen 没有有效 Date")

    if source_date.day <= 7:
        previous_month = (source_date - pd.offsets.MonthEnd(1)).normalize()
        candidates = dates[dates <= previous_month]
        if candidates.empty:
            raise ValueError(f"{source_date.date()} 之前没有可用 screen 月末")
        target = candidates.max()
        if target.to_period("M") != previous_month.to_period("M"):
            raise ValueError(f"月初日期 {source_date.date()} 找不到上月 screen 月末")
        return pd.Timestamp(target)

    same_month = dates[dates.to_period("M") == source_date.to_period("M")]
    if same_month.empty:
        raise ValueError(f"月末日期 {source_date.date()} 在 screen 中找不到同月月末")
    distances = abs(same_month - source_date)
    target = same_month[int(distances.argmin())]
    if abs((target - source_date).days) > 7:
        raise ValueError(f"{source_date.date()} 与最近 screen 月末 {target.date()} 相差超过 7 天")
    return pd.Timestamp(target)


def _find_mapping_header(sheet: Any) -> tuple[int, dict[str, int]]:
    max_header_row = min(sheet.max_row or 1, 20)
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_header_row, values_only=True), 1):
        values = [_clean_text(value) for value in row]
        if "Bloomberg Security" in values:
            return row_number, {value: index + 1 for index, value in enumerate(values) if value}
    raise ValueError("Mapping sheet 中找不到 Bloomberg Security 表头")


def load_isin_mapping(path: Path) -> tuple[dict[str, str], dict[str, Any]]:
    """Read cached Bloomberg values saved by Excel, with manual override priority."""

    formulas_book = load_workbook(path, read_only=True, data_only=False)
    values_book = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Mapping" not in formulas_book.sheetnames or "Mapping" not in values_book.sheetnames:
            raise ValueError("mapping workbook 缺少 Mapping sheet")
        formula_sheet = formulas_book["Mapping"]
        value_sheet = values_book["Mapping"]
        header_row, columns = _find_mapping_header(formula_sheet)
        required = {"Bloomberg Security", "Bloomberg ISIN", "Manual ISIN Override"}
        missing = sorted(required - set(columns))
        if missing:
            raise ValueError(f"mapping workbook 缺少列: {missing}")

        mapping: dict[str, str] = {}
        invalid: list[dict[str, str]] = []
        data_rows = 0
        for values in value_sheet.iter_rows(min_row=header_row + 1, values_only=True):
            security = _clean_text(values[columns["Bloomberg Security"] - 1])
            if not security:
                continue
            data_rows += 1
            manual = _clean_text(values[columns["Manual ISIN Override"] - 1]).upper()
            bloomberg = _clean_text(values[columns["Bloomberg ISIN"] - 1]).upper()
            isin = manual or bloomberg
            if not ISIN_PATTERN.fullmatch(isin):
                invalid.append({"security": security, "value": isin})
                continue
            mapping[security] = isin
        return mapping, {"rows": data_rows, "invalid": invalid}
    finally:
        formulas_book.close()
        values_book.close()


def _recompute_ml_universes(screen: pd.DataFrame, target_dates: set[pd.Timestamp]) -> None:
    mask = screen["Date"].isin(target_dates)
    subset = screen.loc[mask]
    world = subset["Weight in MSCI WORLD"].where(subset["Weight in MSCI WORLD"] > 0)
    sp500 = subset["Weight in SP500"].where(subset["Weight in SP500"] > 0)
    stoxx = subset["Weight in STOXX EUROPE 600"].where(subset["Weight in STOXX EUROPE 600"] > 0)
    us = world.notna() & subset["Exchange Country Name"].eq("UNITED STATES")
    eu = world.notna() & subset["Exchange Country Region"].eq("West Europe")

    replacements = {
        "Weight in Univ ML EU": stoxx.combine_first(world.where(eu)),
        "Weight in Univ ML US": sp500.combine_first(world.where(us)),
        "Weight in Univ ML OTHER": world.where(~us & ~eu),
    }
    for column, values in replacements.items():
        totals = values.groupby(subset["Date"]).transform("sum")
        totals = totals.where(totals.ne(0))
        normalized = pd.to_numeric(values / totals, errors="coerce")
        screen.loc[mask, column] = normalized.to_numpy(dtype="float64", na_value=float("nan"))


def _normalize_benchmark_weights(screen: pd.DataFrame) -> dict[str, Any]:
    report: dict[str, Any] = {}
    for column in BENCHMARK_COLUMNS.values():
        if column not in screen.columns:
            continue
        values = pd.to_numeric(screen[column], errors="coerce")
        positive = values.where(values > 0)
        totals = positive.groupby(screen["Date"]).transform("sum")
        valid = positive.notna() & totals.gt(0)
        before = positive.groupby(screen["Date"]).sum(min_count=1).dropna()
        screen.loc[valid, column] = (positive.loc[valid] / totals.loc[valid] * 100.0).to_numpy(dtype="float64")
        after_values = pd.to_numeric(screen[column], errors="coerce").where(lambda item: item > 0)
        after = after_values.groupby(screen["Date"]).sum(min_count=1).dropna()
        report[column] = {
            "dates_with_positive_weights": int(len(after)),
            "dates_rescaled": int((before.sub(100).abs() > 1e-6).sum()),
            "sum_before_min": float(before.min()),
            "sum_before_max": float(before.max()),
            "sum_after_min": float(after.min()),
            "sum_after_max": float(after.max()),
        }
    return report


def apply_benchmark_weights(
    screen_df: pd.DataFrame,
    snapshots: Iterable[Snapshot],
    isin_mapping: dict[str, str],
) -> tuple[pd.DataFrame, dict[str, Any]]:
    screen = _ensure_isin_column(screen_df)
    screen["ISIN"] = screen["ISIN"].astype("string")
    screen["Date"] = pd.to_datetime(screen["Date"], errors="raise").dt.normalize()
    screen_dates = pd.DatetimeIndex(screen["Date"].unique())
    rows_before = len(screen)

    snapshot_reports: list[dict[str, Any]] = []
    ml_dates: set[pd.Timestamp] = set()
    dropped_unmapped: set[str] = set()
    dropped_missing_screen_key: set[str] = set()
    dropped_unmapped_occurrences = 0
    dropped_missing_screen_key_occurrences = 0
    date_groups = {
        pd.Timestamp(date): pd.Index(indexes)
        for date, indexes in screen.groupby("Date", sort=False).groups.items()
    }
    row_lookups: dict[pd.Timestamp, pd.Series] = {}

    for snapshot in sorted(snapshots, key=lambda item: (item.benchmark, item.source_date, item.source_path.name)):
        target_date = map_source_date(snapshot.source_date, screen_dates)
        column = BENCHMARK_COLUMNS[snapshot.benchmark]
        target_index = date_groups[target_date]
        if target_date not in row_lookups:
            target_isins = screen.loc[target_index, "ISIN"].astype(str)
            if target_isins.duplicated().any():
                duplicates = target_isins[target_isins.duplicated(keep=False)].unique().tolist()
                raise ValueError(f"{target_date.date()} 存在重复 ISIN key，样例: {duplicates[:10]}")
            row_lookups[target_date] = pd.Series(target_index.to_numpy(), index=target_isins.to_numpy())
        row_lookup = row_lookups[target_date]
        target_isins = set(row_lookup.index)

        payload_rows: list[tuple[str, float]] = []
        source_weight = float(sum(row.weight for row in snapshot.rows))
        dropped_unmapped_weight = 0.0
        dropped_missing_key_weight = 0.0
        snapshot_unmapped = 0
        snapshot_missing_key = 0
        for row in snapshot.rows:
            isin = isin_mapping.get(row.security)
            if not isin:
                dropped_unmapped.add(row.security)
                dropped_unmapped_occurrences += 1
                snapshot_unmapped += 1
                dropped_unmapped_weight += row.weight
                continue
            if isin not in target_isins:
                dropped_missing_screen_key.add(f"{row.security}|{isin}")
                dropped_missing_screen_key_occurrences += 1
                snapshot_missing_key += 1
                dropped_missing_key_weight += row.weight
                continue
            payload_rows.append((isin, row.weight))

        if not payload_rows:
            raise ValueError(
                f"{snapshot.benchmark} {snapshot.source_date.date()} 剔除无 mapping/key 成分后没有可写入权重"
            )
        payload = pd.DataFrame(payload_rows, columns=["ISIN", "weight"]).groupby("ISIN", as_index=False)["weight"].sum()
        retained_weight = float(payload["weight"].sum())
        payload["weight"] = payload["weight"] / retained_weight * 100.0

        screen.loc[target_index, column] = float("nan")
        target_rows = payload["ISIN"].map(row_lookup)
        screen.loc[target_rows.to_numpy(), column] = payload["weight"].to_numpy()
        if snapshot.benchmark in {"SP500", "stoxx600"}:
            ml_dates.add(target_date)

        snapshot_reports.append(
            {
                "benchmark": snapshot.benchmark,
                "source_file": str(snapshot.source_path),
                "source_date": str(snapshot.source_date.date()),
                "target_date": str(target_date.date()),
                "source_rows": len(snapshot.rows),
                "retained_source_rows": len(payload_rows),
                "mapped_isins": int(len(payload)),
                "dropped_unmapped_rows": snapshot_unmapped,
                "dropped_missing_screen_key_rows": snapshot_missing_key,
                "source_weight": source_weight,
                "retained_source_weight": retained_weight,
                "dropped_unmapped_weight": float(dropped_unmapped_weight),
                "dropped_missing_screen_key_weight": float(dropped_missing_key_weight),
                "retained_weight_ratio": retained_weight / source_weight,
                "weight_sum": float(screen.loc[target_index, column].sum(skipna=True)),
            }
        )

    normalization_report = _normalize_benchmark_weights(screen)

    if ml_dates:
        _recompute_ml_universes(screen, ml_dates)

    screen = screen.set_index("ISIN")
    contract = validate_screen_contract(screen)
    if not contract["ok"]:
        raise ValueError(f"更新后的 screen 数据契约校验失败: {contract['issues']}")
    if len(screen) != rows_before:
        raise AssertionError(f"screen 行数变化: {rows_before} -> {len(screen)}")

    retained_ratios = [item["retained_weight_ratio"] for item in snapshot_reports]
    return screen, {
        "rows_before": int(rows_before),
        "rows_after": int(len(screen)),
        "snapshots_applied": int(len(snapshot_reports)),
        "snapshot_reports": snapshot_reports,
        "ml_universe_dates_recomputed": int(len(ml_dates)),
        "benchmark_normalization": normalization_report,
        "dropped_unmapped_occurrences": dropped_unmapped_occurrences,
        "dropped_unmapped_unique_securities": sorted(dropped_unmapped),
        "dropped_missing_screen_key_occurrences": dropped_missing_screen_key_occurrences,
        "dropped_missing_screen_key_unique_securities": sorted(dropped_missing_screen_key),
        "retained_weight_ratio_min": float(min(retained_ratios)),
        "retained_weight_ratio_mean": float(sum(retained_ratios) / len(retained_ratios)),
    }


def _timestamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _refresh_derived_outputs(screen: pd.DataFrame) -> dict[str, Any]:
    latest_date = pd.to_datetime(screen["Date"]).max()
    latest = screen.loc[pd.to_datetime(screen["Date"]).eq(latest_date)].copy()
    five_year_start = latest_date - pd.DateOffset(years=5)
    five_year = screen.loc[pd.to_datetime(screen["Date"]).ge(five_year_start)].copy()
    latest.to_parquet(LAST_SCREEN_PATH, index=True)
    five_year.to_parquet(SCREEN_5Y_PATH, index=True)
    return {"latest_date": str(latest_date.date()), "last_screen_rows": int(len(latest)), "screen_5y_rows": int(len(five_year))}


def _archive_inputs(stamp: str) -> str:
    archive_root = INCOMING_ROOT.parent / "archive" / "benchmark_weights" / stamp
    archive_root.mkdir(parents=True, exist_ok=False)
    for benchmark in BENCHMARK_COLUMNS:
        source = INCOMING_ROOT / benchmark
        if source.exists():
            shutil.move(str(source), str(archive_root / benchmark))
    return str(archive_root)


def run(
    mapping_workbook: Path,
    *,
    write: bool,
    archive_inputs: bool,
    incoming_root: Path = INCOMING_ROOT,
) -> dict[str, Any]:
    stamp = _timestamp()
    snapshots, skipped = collect_snapshots(incoming_root)
    mapping, mapping_report = load_isin_mapping(mapping_workbook)
    required_securities = {row.security for snapshot in snapshots for row in snapshot.rows}
    missing_mapping = sorted(required_securities - set(mapping))

    original = pd.read_parquet(SCREEN_PATH)
    updated, apply_report = apply_benchmark_weights(original, snapshots, mapping)
    report: dict[str, Any] = {
        "status": "dry_run_ok" if not write else "success",
        "mapping_workbook": str(mapping_workbook),
        "incoming_root": str(incoming_root),
        "screen_path": str(SCREEN_PATH),
        "source_snapshots": len(snapshots),
        "skipped_non_point_files": skipped,
        "required_securities": len(required_securities),
        "unmapped_securities_dropped": missing_mapping,
        "mapping_report": mapping_report,
        "apply_report": apply_report,
        "write": write,
    }

    if write:
        backup_dir = SCREEN_PATH.parent / "backups" / SCREEN_PATH.stem
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{SCREEN_PATH.stem}_{stamp}_before_benchmark_weight_import{SCREEN_PATH.suffix}"
        shutil.copy2(SCREEN_PATH, backup_path)
        temp_path = SCREEN_PATH.with_name(f"{SCREEN_PATH.stem}.{stamp}.benchmark_import_tmp{SCREEN_PATH.suffix}")
        updated.to_parquet(temp_path, index=True)
        check = pd.read_parquet(temp_path, columns=["Date", *BENCHMARK_COLUMNS.values()])
        if len(check) != len(updated):
            raise AssertionError("临时 parquet 行数校验失败")
        temp_path.replace(SCREEN_PATH)
        report["backup_path"] = str(backup_path)
        report["derived_outputs"] = _refresh_derived_outputs(updated)
        if archive_inputs:
            report["archive_path"] = _archive_inputs(stamp)

    qa_path = QA_ROOT / f"benchmark_weight_import_{stamp}.json"
    manifest_path = MANIFEST_ROOT / f"benchmark_weight_import_{stamp}.json"
    _write_json(qa_path, report)
    _write_json(manifest_path, report)
    report["qa_path"] = str(qa_path)
    report["manifest_path"] = str(manifest_path)
    return report


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="回补 Bloomberg 历史 benchmark 权重到 canonical screen")
    parser.add_argument("--mapping-workbook", required=True, type=Path)
    parser.add_argument("--incoming-root", type=Path, default=INCOMING_ROOT)
    parser.add_argument("--write", action="store_true", help="实际写入；默认只执行 dry-run")
    parser.add_argument("--no-archive-inputs", action="store_true", help="成功写入后保留 incoming benchmark 目录")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    report = run(
        args.mapping_workbook,
        write=args.write,
        archive_inputs=args.write and not args.no_archive_inputs,
        incoming_root=args.incoming_root,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
