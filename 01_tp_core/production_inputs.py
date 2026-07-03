"""整理月度生产数据更新使用的原始输入文件。

整理逻辑坚持“内容优先”：历史文件可能存在误导性的文件名或扩展名，
因此先读取 schema 与日期字段，再判断类别、期间和标准化归档名称。
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import re
import shutil
from pathlib import Path
from typing import Any

import pandas as pd

from .data_contract import (
    DATE_COLUMN,
    RISK_COLUMNS,
    SEDOL_COLUMN,
    deprecated_screen_columns,
    sedol_coverage,
    validate_returns_contract,
    validate_screen_contract,
    weight_sum_report,
)
from .data_sources import PRODUCTION_INPUTS_DIR, RETURNS_PATH, SCREEN_AGGREGATE_PATH, SCREEN_DIR, TP_ROOT

SCREEN_REQUIRED_HEADERS = {"Symbol", "ISIN", "Date"}
CIQ_REQUIRED_COLUMNS = {"ISIN", "Date"}


def _now_stamp() -> str:
    return dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _iso_from_timestamp(value: float) -> str:
    return dt.datetime.fromtimestamp(value).isoformat(timespec="seconds")


def _json_default(value: Any) -> str | int | float | None:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (pd.Timestamp, dt.datetime, dt.date)):
        if pd.isna(value):
            return None
        return value.isoformat()
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    if pd.isna(value):
        return None
    return str(value)


def _safe_relpath(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except ValueError:
        return str(path)


def _sanitize_token(value: str, fallback: str = "source") -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return token[:80] or fallback


def _source_id(path: Path, root: Path) -> str:
    stat = path.stat()
    payload = f"{_safe_relpath(path, root)}|{stat.st_size}|{stat.st_mtime_ns}".encode("utf-8")
    return hashlib.sha1(payload).hexdigest()[:10]


def _parse_any_date(value: Any) -> pd.Timestamp | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed).normalize()


def _month_end(value: pd.Timestamp) -> pd.Timestamp:
    return pd.Timestamp(value).to_period("M").to_timestamp("M")


def _format_date(value: pd.Timestamp | None) -> str | None:
    if value is None or pd.isna(value):
        return None
    return pd.Timestamp(value).strftime("%Y-%m-%d")


def _format_ym(value: pd.Timestamp | None) -> str:
    if value is None or pd.isna(value):
        return "unknown"
    return pd.Timestamp(value).strftime("%Y%m")


def _format_yyyymmdd(value: pd.Timestamp | None) -> str:
    if value is None or pd.isna(value):
        return "unknown"
    return pd.Timestamp(value).strftime("%Y%m%d")


def _parse_filename_date(stem: str) -> dict[str, Any] | None:
    clean = stem.strip()
    compact = re.sub(r"[^0-9]", "", clean)

    if re.fullmatch(r"\d{8}", compact):
        parsed = _parse_any_date(f"{compact[:4]}-{compact[4:6]}-{compact[6:8]}")
        if parsed is not None:
            return {"date": parsed, "granularity": "day", "token": compact}

    if re.fullmatch(r"\d{6}", compact):
        parsed = _parse_any_date(f"{compact[:4]}-{compact[4:6]}-01")
        if parsed is not None:
            return {"date": _month_end(parsed), "granularity": "month", "token": compact}

    if re.fullmatch(r"\d{4}", compact):
        yy = int(compact[:2])
        year = 2000 + yy if yy < 80 else 1900 + yy
        parsed = _parse_any_date(f"{year:04d}-{compact[2:4]}-01")
        if parsed is not None:
            return {"date": _month_end(parsed), "granularity": "month", "token": compact}

    match = re.search(r"(19\d{2}|20\d{2})[-_ ](\d{2})(?:[-_ ](\d{2}))?", clean)
    if match:
        year, month, day = match.groups()
        day = day or "01"
        parsed = _parse_any_date(f"{year}-{month}-{day}")
        if parsed is not None:
            granularity = "day" if match.group(3) else "month"
            return {
                "date": parsed if granularity == "day" else _month_end(parsed),
                "granularity": granularity,
                "token": match.group(0),
            }

    return None


def _name_relation(name_date: pd.Timestamp | None, date_min: pd.Timestamp | None, date_max: pd.Timestamp | None) -> str:
    if name_date is None:
        return "unparseable"
    if date_min is None or date_max is None:
        return "no_content_date"
    if date_min <= name_date <= date_max:
        return "inside_content_range"
    if name_date.to_period("M") in {date_min.to_period("M"), date_max.to_period("M")}:
        return "same_boundary_month"
    delta_to_max = int((name_date - date_max).days)
    return f"outside_content_range_delta_to_max_{delta_to_max}d"


def _monthly_name_relation(name_date: pd.Timestamp | None, production_month_end: pd.Timestamp | None) -> str:
    if name_date is None:
        return "unparseable"
    if production_month_end is None:
        return "no_content_date"
    if name_date.to_period("M") == production_month_end.to_period("M"):
        return "same_content_month"
    delta = int((name_date - production_month_end).days)
    return f"different_content_month_delta_{delta}d"


def _base_record(path: Path, root: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "source_path": str(path),
        "source_relpath": _safe_relpath(path, root),
        "source_name": path.name,
        "source_stem": path.stem,
        "source_suffix": path.suffix,
        "source_size_bytes": int(stat.st_size),
        "source_mtime": _iso_from_timestamp(stat.st_mtime),
        "source_id": _source_id(path, root),
        "status": "skipped",
        "reason": None,
    }


def inspect_screen_excel(path: Path, root: Path) -> dict[str, Any]:
    record = _base_record(path, root)
    record.update({"source_group": "screen_monthly", "detected_format": "xlsx"})

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=5, max_row=5, values_only=True))
        header_values = [str(value).strip() if value is not None else "" for value in header]
        header_set = {value for value in header_values if value}
        date_col = next((idx + 1 for idx, value in enumerate(header_values) if value == "Date"), None)

        record.update(
            {
                "sheet_name": worksheet.title,
                "rows_estimate": max(int(worksheet.max_row or 0) - 6, 0),
                "columns": int(worksheet.max_column or 0),
                "columns_sample": header_values[:20],
            }
        )

        if not SCREEN_REQUIRED_HEADERS.issubset(header_set) or date_col is None:
            workbook.close()
            record.update({"reason": "not_factset_screen_export"})
            return record

        dates: list[pd.Timestamp] = []
        max_scan_row = min(int(worksheet.max_row or 0), 206)
        for row in worksheet.iter_rows(
            min_row=7,
            max_row=max_scan_row,
            min_col=date_col,
            max_col=date_col,
            values_only=True,
        ):
            parsed = _parse_any_date(row[0])
            if parsed is not None:
                dates.append(parsed)
        workbook.close()

        if not dates:
            record.update({"reason": "screen_date_not_found"})
            return record

        date_min = min(dates)
        date_max = max(dates)
        production_month_end = _month_end(date_max)
        name_info = _parse_filename_date(path.stem)
        name_date = name_info["date"] if name_info else None
        relation = _monthly_name_relation(name_date, production_month_end)

        record.update(
            {
                "status": "eligible",
                "category": "screen_monthly",
                "reason": "content_schema_match",
                "date_min": _format_date(date_min),
                "date_max": _format_date(date_max),
                "production_month_end": _format_date(production_month_end),
                "date_source": "excel_date_column_sample",
                "name_inferred_date": _format_date(name_date),
                "name_inferred_granularity": name_info["granularity"] if name_info else None,
                "name_content_relation": relation,
                "name_mismatch": relation.startswith("different_content_month"),
                "extension_mismatch": path.suffix.lower() != ".xlsx",
                "normalized_name": f"screen_monthly_{_format_yyyymmdd(production_month_end)}.xlsx",
                "archive_bucket": _format_ym(production_month_end),
            }
        )
        return record
    except Exception as exc:
        record.update({"status": "error", "reason": f"excel_read_error: {type(exc).__name__}: {exc}"})
        return record


def inspect_returns_parquet(path: Path, root: Path) -> dict[str, Any]:
    record = _base_record(path, root)
    record.update({"source_group": "returns_delta", "detected_format": "parquet_candidate"})
    try:
        frame = pd.read_parquet(path)
    except Exception as exc:
        record.update({"reason": f"not_parquet: {type(exc).__name__}: {str(exc)[:180]}"})
        return record

    dates = pd.to_datetime(frame.index, errors="coerce")
    valid_dates = dates[~pd.isna(dates)]
    if len(valid_dates) == 0 or len(frame.columns) < 100:
        record.update(
            {
                "reason": "parquet_but_not_returns_delta",
                "rows": int(len(frame)),
                "columns": int(len(frame.columns)),
                "columns_sample": [str(column) for column in frame.columns[:20]],
            }
        )
        return record

    date_min = pd.Timestamp(valid_dates.min()).normalize()
    date_max = pd.Timestamp(valid_dates.max()).normalize()
    name_info = _parse_filename_date(path.stem)
    name_date = name_info["date"] if name_info else None
    relation = _name_relation(name_date, date_min, date_max)

    record.update(
        {
            "status": "eligible",
            "category": "returns_delta",
            "reason": "content_schema_match",
            "rows": int(len(frame)),
            "columns": int(len(frame.columns)),
            "columns_sample": [str(column) for column in frame.columns[:20]],
            "date_min": _format_date(date_min),
            "date_max": _format_date(date_max),
            "date_source": "parquet_index",
            "name_inferred_date": _format_date(name_date),
            "name_inferred_granularity": name_info["granularity"] if name_info else None,
            "name_content_relation": relation,
            "name_mismatch": relation.startswith("outside_content_range"),
            "extension_mismatch": path.suffix.lower() != ".parquet",
            "normalized_name": f"returns_delta_{_format_yyyymmdd(date_min)}_{_format_yyyymmdd(date_max)}.parquet",
            "archive_bucket": _format_ym(date_max),
        }
    )
    return record


def inspect_ciq_parquet(path: Path, root: Path) -> dict[str, Any]:
    record = _base_record(path, root)
    record.update({"source_group": "ciq_history", "detected_format": "parquet_candidate"})
    try:
        frame = pd.read_parquet(path)
    except Exception as exc:
        record.update({"reason": f"not_parquet: {type(exc).__name__}: {str(exc)[:180]}"})
        return record

    columns = {str(column) for column in frame.columns}
    if not CIQ_REQUIRED_COLUMNS.issubset(columns):
        record.update(
            {
                "reason": "parquet_but_not_ciq",
                "rows": int(len(frame)),
                "columns": int(len(frame.columns)),
                "columns_sample": [str(column) for column in frame.columns[:20]],
            }
        )
        return record

    dates = pd.to_datetime(frame["Date"], errors="coerce")
    valid_dates = dates[~pd.isna(dates)]
    if len(valid_dates) == 0:
        record.update({"reason": "ciq_date_not_found"})
        return record

    date_min = pd.Timestamp(valid_dates.min()).normalize()
    date_max = pd.Timestamp(valid_dates.max()).normalize()
    name_info = _parse_filename_date(path.stem)
    name_date = name_info["date"] if name_info else None
    relation = _name_relation(name_date, date_min, date_max)

    record.update(
        {
            "status": "eligible",
            "category": "ciq_history",
            "reason": "content_schema_match",
            "rows": int(len(frame)),
            "columns": int(len(frame.columns)),
            "columns_sample": [str(column) for column in frame.columns[:20]],
            "date_min": _format_date(date_min),
            "date_max": _format_date(date_max),
            "date_source": "parquet_date_column",
            "name_inferred_date": _format_date(name_date),
            "name_inferred_granularity": name_info["granularity"] if name_info else None,
            "name_content_relation": relation,
            "name_mismatch": relation.startswith("outside_content_range"),
            "extension_mismatch": path.suffix.lower() != ".parquet",
            "normalized_name": f"ciq_history_{_format_yyyymmdd(date_min)}_{_format_yyyymmdd(date_max)}.parquet",
            "archive_bucket": f"{_format_yyyymmdd(date_min)}_{_format_yyyymmdd(date_max)}",
        }
    )
    return record


def _collect_incoming_inventory(root: Path, records: list[dict[str, Any]]) -> None:
    screen_dir = SCREEN_DIR if SCREEN_DIR.is_absolute() else root / SCREEN_DIR
    incoming_dir = screen_dir / "production_inputs" / "incoming"
    if not incoming_dir.exists():
        return

    for batch_dir in sorted(path for path in incoming_dir.iterdir() if path.is_dir()):
        screen_dir = batch_dir / "screen"
        if screen_dir.exists():
            for path in sorted(screen_dir.glob("*.xlsx")):
                if not path.name.startswith("~$"):
                    records.append(inspect_screen_excel(path, root))

        returns_dir = batch_dir / "returns"
        if returns_dir.exists():
            for path in sorted(item for item in returns_dir.iterdir() if item.is_file()):
                records.append(inspect_returns_parquet(path, root))

        ciq_dir = batch_dir / "ciq"
        if ciq_dir.exists():
            for path in sorted(item for item in ciq_dir.iterdir() if item.is_file()):
                records.append(inspect_ciq_parquet(path, root))


def _collect_legacy_inventory(root: Path, records: list[dict[str, Any]]) -> None:
    screen_dir = SCREEN_DIR if SCREEN_DIR.is_absolute() else root / SCREEN_DIR

    monthly_dir = screen_dir / "monthly"
    monthly_candidates = sorted(monthly_dir.glob("*.xlsx")) if monthly_dir.exists() else []
    monthly_old_dir = monthly_dir / "old"
    if monthly_old_dir.exists():
        monthly_candidates.extend(sorted(monthly_old_dir.glob("*.xlsx")))
    for path in monthly_candidates:
        if path.name.startswith("~$"):
            continue
        records.append(inspect_screen_excel(path, root))

    returns_dir = screen_dir / "returns"
    if returns_dir.exists():
        for path in sorted(item for item in returns_dir.iterdir() if item.is_file()):
            records.append(inspect_returns_parquet(path, root))

    ciq_dir = screen_dir / "ciq"
    if ciq_dir.exists():
        for path in sorted(item for item in ciq_dir.rglob("*") if item.is_file()):
            records.append(inspect_ciq_parquet(path, root))


def collect_inventory(root: Path, include_legacy_sources: bool = False) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    _collect_incoming_inventory(root, records)
    if include_legacy_sources:
        _collect_legacy_inventory(root, records)
    return records


def _target_for_record(record: dict[str, Any], production_dir: Path) -> Path | None:
    category = record.get("category")
    normalized_name = record.get("normalized_name")
    if not category or not normalized_name:
        return None
    if category == "screen_monthly":
        return production_dir / "archive" / "screen_monthly" / str(record.get("archive_bucket", "unknown")) / normalized_name
    if category == "returns_delta":
        return production_dir / "archive" / "returns_delta" / str(record.get("archive_bucket", "unknown")) / normalized_name
    if category == "ciq_history":
        return production_dir / "archive" / "ciq_history" / str(record.get("archive_bucket", "unknown")) / normalized_name
    return production_dir / "archive" / "unclassified" / normalized_name


def _resolve_collision(target: Path, record: dict[str, Any], used_targets: set[Path]) -> Path:
    if target not in used_targets and not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    source_token = _sanitize_token(str(record.get("source_stem") or "source"))
    candidate = target.with_name(f"{stem}_source-{source_token}_{record['source_id']}{suffix}")
    counter = 2
    while candidate in used_targets or candidate.exists():
        candidate = target.with_name(f"{stem}_source-{source_token}_{record['source_id']}_{counter}{suffix}")
        counter += 1
    return candidate


def _dedupe_key(record: dict[str, Any]) -> tuple[str, str] | None:
    if record.get("status") != "eligible":
        return None
    category = record.get("category")
    normalized_name = record.get("normalized_name")
    archive_bucket = record.get("archive_bucket")
    if not category or not normalized_name:
        return None
    return (str(category), f"{archive_bucket}/{normalized_name}")


def _canonical_rank(record: dict[str, Any]) -> tuple[int, int, int, str]:
    columns = int(record.get("columns") or 0)
    rows = int(record.get("rows") or record.get("rows_estimate") or 0)
    size = int(record.get("source_size_bytes") or 0)
    mtime = str(record.get("source_mtime") or "")
    return (columns, rows, size, mtime)


def _canonical_records(records: list[dict[str, Any]], keep_duplicates: bool) -> dict[int, dict[str, Any]]:
    if keep_duplicates:
        return {id(record): record for record in records if record.get("status") == "eligible"}

    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for record in records:
        key = _dedupe_key(record)
        if key is not None:
            groups.setdefault(key, []).append(record)

    canonical_by_id: dict[int, dict[str, Any]] = {}
    for group_records in groups.values():
        group_records.sort(key=_canonical_rank, reverse=True)
        canonical = group_records[0]
        canonical["canonical_source_for_period"] = True
        canonical["duplicate_group_size"] = len(group_records)
        canonical_by_id[id(canonical)] = canonical
        for duplicate in group_records[1:]:
            duplicate["canonical_source_for_period"] = False
            duplicate["duplicate_group_size"] = len(group_records)
            duplicate["duplicate_of_source_relpath"] = canonical.get("source_relpath")
            duplicate["duplicate_of_normalized_name"] = canonical.get("normalized_name")
    return canonical_by_id


def materialize_archive(
    records: list[dict[str, Any]],
    production_dir: Path,
    copy_files: bool,
    keep_duplicates: bool = False,
) -> None:
    canonical_by_id = _canonical_records(records, keep_duplicates=keep_duplicates)
    used_targets: set[Path] = set()
    copied_targets_by_record_id: dict[int, Path] = {}

    for record in records:
        if record.get("status") != "eligible":
            record["archive_action"] = "none"
            continue

        if id(record) not in canonical_by_id:
            record["archive_action"] = "duplicate_not_copied"
            continue

        source = Path(str(record["source_path"]))
        target = _target_for_record(record, production_dir)
        if target is None:
            record["archive_action"] = "none"
            continue

        target = _resolve_collision(target, record, used_targets)
        used_targets.add(target)
        copied_targets_by_record_id[id(record)] = target
        record["target_path"] = str(target)
        record["target_relpath"] = _safe_relpath(target, production_dir)

        if not copy_files:
            record["archive_action"] = "dry_run"
            continue

        target.parent.mkdir(parents=True, exist_ok=True)
        if target.exists() and target.stat().st_size == source.stat().st_size:
            record["archive_action"] = "already_exists_same_size"
            continue
        shutil.copy2(source, target)
        record["archive_action"] = "copied"

    canonical_target_by_source = {
        record.get("source_relpath"): record.get("target_relpath")
        for record in records
        if record.get("canonical_source_for_period") and record.get("target_relpath")
    }
    for record in records:
        if record.get("archive_action") == "duplicate_not_copied":
            record["duplicate_of_target_relpath"] = canonical_target_by_source.get(
                record.get("duplicate_of_source_relpath")
            )


def _write_csv(records: list[dict[str, Any]], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "status",
        "category",
        "source_relpath",
        "source_name",
        "source_suffix",
        "source_size_bytes",
        "date_min",
        "date_max",
        "production_month_end",
        "rows",
        "columns",
        "normalized_name",
        "target_relpath",
        "archive_action",
        "name_inferred_date",
        "name_content_relation",
        "name_mismatch",
        "extension_mismatch",
        "reason",
        "canonical_source_for_period",
        "duplicate_group_size",
        "duplicate_of_source_relpath",
        "duplicate_of_target_relpath",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow(record)


def _records_summary(records: list[dict[str, Any]]) -> dict[str, Any]:
    summary: dict[str, Any] = {
        "total_records": len(records),
        "eligible_records": sum(1 for record in records if record.get("status") == "eligible"),
        "skipped_records": sum(1 for record in records if record.get("status") == "skipped"),
        "error_records": sum(1 for record in records if record.get("status") == "error"),
        "copied_records": sum(1 for record in records if record.get("archive_action") == "copied"),
        "already_existing_records": sum(1 for record in records if record.get("archive_action") == "already_exists_same_size"),
        "duplicate_not_copied_records": sum(
            1 for record in records if record.get("archive_action") == "duplicate_not_copied"
        ),
        "name_mismatch_records": sum(1 for record in records if record.get("name_mismatch")),
        "extension_mismatch_records": sum(1 for record in records if record.get("extension_mismatch")),
        "by_category": {},
        "bytes_by_category": {},
    }
    for record in records:
        category = record.get("category") or f"skipped:{record.get('source_group', 'unknown')}"
        summary["by_category"][category] = summary["by_category"].get(category, 0) + 1
        if record.get("status") == "eligible":
            summary["bytes_by_category"][category] = summary["bytes_by_category"].get(category, 0) + int(
                record.get("source_size_bytes") or 0
            )
    return summary


def write_manifest(
    records: list[dict[str, Any]],
    production_dir: Path,
    stamp: str,
    copy_files: bool,
    keep_duplicates: bool,
    include_legacy_sources: bool,
) -> dict[str, Path]:
    manifest_dir = production_dir / "manifests"
    manifest_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = manifest_dir / f"input_inventory_{stamp}.json"
    latest_path = manifest_dir / "input_inventory_latest.json"
    csv_path = manifest_dir / f"input_inventory_{stamp}.csv"

    payload = {
        "run_id": stamp,
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "mode": "copy" if copy_files else "dry_run",
        "source_scope": "incoming_plus_legacy" if include_legacy_sources else "incoming_only",
        "dedupe_policy": "keep_duplicates" if keep_duplicates else "one_canonical_copy_per_content_period",
        "production_inputs_dir": str(production_dir),
        "summary": _records_summary(records),
        "records": records,
    }
    manifest_path.write_text(json.dumps(payload, indent=2, default=_json_default), encoding="utf-8")
    latest_path.write_text(json.dumps(payload, indent=2, default=_json_default), encoding="utf-8")
    _write_csv(records, csv_path)
    return {"json": manifest_path, "latest_json": latest_path, "csv": csv_path}


def generate_database_profile(root: Path, production_dir: Path, stamp: str) -> dict[str, Path] | None:
    profile_dir = production_dir / "profiles"
    profile_dir.mkdir(parents=True, exist_ok=True)
    profile_path = profile_dir / f"latest_database_profile_{stamp}.json"
    latest_path = profile_dir / "latest_database_profile_latest.json"

    screen_path = SCREEN_AGGREGATE_PATH if SCREEN_AGGREGATE_PATH.is_absolute() else root / SCREEN_AGGREGATE_PATH
    returns_path = RETURNS_PATH if RETURNS_PATH.is_absolute() else root / RETURNS_PATH
    profile: dict[str, Any] = {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "screen_aggregate_path": str(screen_path),
        "returns_path": str(returns_path),
        "screen_aggregate_exists": screen_path.exists(),
        "returns_exists": returns_path.exists(),
    }

    screen_df: pd.DataFrame | None = None
    returns_df: pd.DataFrame | None = None

    if screen_path.exists():
        screen_df = pd.read_parquet(screen_path)
        screen_contract = validate_screen_contract(screen_df)
        screen_work = screen_df.reset_index() if "ISIN" not in screen_df.columns and screen_df.index.name == "ISIN" else screen_df.copy()
        screen_work[DATE_COLUMN] = pd.to_datetime(screen_work[DATE_COLUMN], errors="coerce")
        latest_date = screen_work[DATE_COLUMN].max()
        latest_slice = screen_work.loc[screen_work[DATE_COLUMN] == latest_date]
        risk_report = {}
        for column in RISK_COLUMNS:
            risk_report[column] = {
                "exists": column in screen_work.columns,
                "non_null_all": int(screen_work[column].notna().sum()) if column in screen_work.columns else 0,
                "non_null_latest": int(latest_slice[column].notna().sum()) if column in screen_work.columns else 0,
            }
        profile["screen_aggregate"] = {
            **screen_contract,
            "file_size_bytes": int(screen_path.stat().st_size),
            "file_mtime": _iso_from_timestamp(screen_path.stat().st_mtime),
            "latest_date": _format_date(latest_date),
            "latest_rows": int(len(latest_slice)),
            "deprecated_columns": deprecated_screen_columns(screen_work.columns),
            "risk_columns": risk_report,
            "weight_sums_latest": weight_sum_report(screen_work),
        }

    if returns_path.exists():
        returns_df = pd.read_parquet(returns_path)
        returns_contract = validate_returns_contract(returns_df)
        returns_numeric = returns_df.select_dtypes(include="number")
        profile["returns"] = {
            **returns_contract,
            "file_size_bytes": int(returns_path.stat().st_size),
            "file_mtime": _iso_from_timestamp(returns_path.stat().st_mtime),
            "max_return": float(returns_numeric.max().max()) if not returns_numeric.empty else None,
            "min_return": float(returns_numeric.min().min()) if not returns_numeric.empty else None,
        }

    if screen_df is not None and returns_df is not None:
        profile["sedol_coverage_latest"] = sedol_coverage(screen_df, returns_df, latest_only=True)

    profile_path.write_text(json.dumps(profile, indent=2, default=_json_default), encoding="utf-8")
    latest_path.write_text(json.dumps(profile, indent=2, default=_json_default), encoding="utf-8")
    return {"json": profile_path, "latest_json": latest_path}


def write_readme(production_dir: Path) -> Path:
    readme = production_dir / "README.md"
    text = r"""# 生产输入目录

此目录是月度 screen 更新使用的原始输入文件的受控入口。历史文件可能存在错误命名或错误扩展名，因此整理流程始终以文件内容为准，而不是以原始文件名为准。

## 目录结构

- `incoming/YYYYMM/screen/`：本月新收到的 FactSet 月度 Screen Excel，运行前放入。
- `incoming/YYYYMM/returns/`：本月新收到的日频 returns parquet 增量，运行前放入。
- `incoming/YYYYMM/ciq/`：本月新收到的 CIQ parquet 增量，运行前放入。
- `archive/screen_monthly/YYYYMM/`：按内容日期标准化后的历史 screen 输入。
- `archive/returns_delta/YYYYMM/`：按日期范围标准化后的 returns 增量。
- `archive/ciq_history/YYYYMMDD_YYYYMMDD/`：按日期范围标准化后的 CIQ 历史分片。
- `manifests/`：输入清单，记录原始路径、内容日期、标准化名称、命名不一致和扩展名不一致标记。
- `profiles/`：当前 canonical `screen_aggregate.parquet` 与 `returns.parquet` 的概况报告。

## 命名规则

- 月度 Screen 输入：`screen_monthly_YYYYMMDD.xlsx`，其中 `YYYYMMDD` 来自 Excel `Date` 列推断出的生产月末。
- Returns 增量：`returns_delta_YYYYMMDD_YYYYMMDD.parquet`，日期范围来自 parquet 的日期索引。
- CIQ 历史：`ciq_history_YYYYMMDD_YYYYMMDD.parquet`，日期范围来自 parquet 的 `Date` 列。

原始文件名只用于审计和 mismatch 检测，不作为主要日期信号。

## 月度流程

1. 将新原始文件放入 `incoming/YYYYMM/screen`、`incoming/YYYYMM/returns` 和 `incoming/YYYYMM/ciq`。
2. 在 `C:\GoogleDrive\TP` 运行 `python -m 01_tp_core.production_inputs`，对 `incoming` 中的可识别文件进行分类和归档。默认每个内容期间只保留一个 canonical 副本；重复来源会记录在清单里，但不会再次复制。只有迁移历史目录时才使用 `--include-legacy-sources`，只有明确需要保留所有重复物理文件时才使用 `--keep-duplicates`。
3. 查看 `manifests/input_inventory_latest.json`，重点检查 `name_mismatch`、`extension_mismatch`、重复来源、跳过文件和复制动作。
4. 月度更新只对 canonical 数据集运行：`00_screen/screen_aggregate.parquet` 与 `00_screen/returns.parquet`。
5. 将生成的 `profiles/latest_database_profile_latest.json` 与当月 QA 记录一起保留。
"""
    readme.write_text(text, encoding="utf-8")
    return readme


def ensure_layout(production_dir: Path) -> None:
    for relative in [
        "incoming",
        "archive/screen_monthly",
        "archive/returns_delta",
        "archive/ciq_history",
        "archive/unclassified",
        "manifests",
        "profiles",
    ]:
        (production_dir / relative).mkdir(parents=True, exist_ok=True)


def run(
    root: Path,
    production_dir: Path,
    copy_files: bool = True,
    profile: bool = True,
    keep_duplicates: bool = False,
    include_legacy_sources: bool = False,
) -> dict[str, Any]:
    stamp = _now_stamp()
    ensure_layout(production_dir)
    write_readme(production_dir)
    records = collect_inventory(root, include_legacy_sources=include_legacy_sources)
    materialize_archive(records, production_dir, copy_files=copy_files, keep_duplicates=keep_duplicates)
    manifest_paths = write_manifest(
        records,
        production_dir,
        stamp,
        copy_files=copy_files,
        keep_duplicates=keep_duplicates,
        include_legacy_sources=include_legacy_sources,
    )
    profile_paths = generate_database_profile(root, production_dir, stamp) if profile else None
    return {
        "run_id": stamp,
        "summary": _records_summary(records),
        "manifest_paths": {key: str(value) for key, value in manifest_paths.items()},
        "profile_paths": {key: str(value) for key, value in profile_paths.items()} if profile_paths else None,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="整理 TP 月度生产输入文件。", add_help=False)
    parser._optionals.title = "选项"
    parser.add_argument("-h", "--help", action="help", help="显示帮助信息并退出。")
    parser.add_argument("--root", type=Path, default=TP_ROOT, help="TP 工作区根目录。")
    parser.add_argument(
        "--production-dir",
        type=Path,
        default=PRODUCTION_INPUTS_DIR,
        help="生产输入归档目标目录。",
    )
    parser.add_argument("--dry-run", action="store_true", help="只生成清单，不复制归档文件。")
    parser.add_argument("--no-profile", action="store_true", help="跳过 canonical 数据集概况报告生成。")
    parser.add_argument(
        "--keep-duplicates",
        action="store_true",
        help="复制重复来源文件；默认每个内容期间只保留一个 canonical 副本。",
    )
    parser.add_argument(
        "--include-legacy-sources",
        action="store_true",
        help="额外扫描旧的 00_screen/monthly、00_screen/returns、00_screen/ciq；默认只扫描 production_inputs/incoming。",
    )
    args = parser.parse_args(argv)

    result = run(
        root=args.root,
        production_dir=args.production_dir,
        copy_files=not args.dry_run,
        profile=not args.no_profile,
        keep_duplicates=args.keep_duplicates,
        include_legacy_sources=args.include_legacy_sources,
    )
    print(json.dumps(result, indent=2, default=_json_default))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
