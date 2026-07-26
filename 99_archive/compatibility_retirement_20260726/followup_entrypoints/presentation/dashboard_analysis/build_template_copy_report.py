from __future__ import annotations

import json
import math
import re
from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from tp_core.workspace import BACKTEST_RUNS_DIR

ROOT = Path(__file__).resolve().parents[3]
APP_ROOT = Path(__file__).resolve().parent
TEMPLATE = APP_ROOT / "analyse.xlsx"
RUN_DIR = BACKTEST_RUNS_DIR / "ad_hoc" / "min_te_score_ml_202505_140_160_20260701_120512"
SEC_LIST_PATH = RUN_DIR / "sec_list_min_te_score_ml_202505_140_160.parquet"
OPT_RESULT_PATH = RUN_DIR / "optimizer_result_min_te_score_ml_202505_140_160.parquet"
SUMMARY_PATH = RUN_DIR / "summary.json"
RETURNS_PATH = ROOT / "00_screen" / "returns.parquet"
OUT_DIR = APP_ROOT / "outputs"
OUT_PATH = OUT_DIR / "analyse_min_te_score_ml_202505_template_copy.xlsx"

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)

ICB_MAP = {
    1: "Auto & Parts", 2: "Banks", 3: "Basic Resources", 4: "Chemicals",
    5: "Construction", 6: "Financial Services", 7: "Food, Beverage & Tobacco",
    8: "Health Care", 9: "Industrial Goods & Services", 10: "Insurance",
    11: "Media", 12: "Energy", 13: "Personal & Household Goods",
    14: "Real Estate", 15: "Retail", 16: "Technology",
    17: "Telecommunications", 18: "Travel & Leisure", 19: "Utilities",
}

ALIASES = {
    "Score Dividend": "Dividend Avg Percentile",
    "Score Value": "Value Avg Percentile",
    "Score Quality": "Quality Avg Percentile",
    "Score Momentum": "Mom Avg Percentile",
    "Score Volatility": "LowVol Avg Percentile",
    "Score Growth": "Growth Avg Percentile",
    "Score Multifacteur": "Multi Avg Percentile",
    "Score Multiffacteur Tilt": "Multi Avg Percentile",
    "Score ML rebased": "Score ML",
    " Benchmark ICB Supersector ": "Secto",
    "Benchmark Market Value Millions in EUR ": "Benchmark Market Value Millions in EUR BK",
}


def template_headers(sheet_name: str) -> list[str]:
    wb = load_workbook(TEMPLATE, read_only=True, data_only=False)
    ws = wb[sheet_name]
    values = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    wb.close()
    return [str(v) for v in values if v is not None]


def col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch.upper()) - 64
    return n


def num_to_col(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(r + 65) + s
    return s


def cell_ref(row: int, col: int) -> str:
    return f"{num_to_col(col)}{row}"


def split_cell(ref: str) -> tuple[int, int]:
    m = re.fullmatch(r"([A-Z]+)([0-9]+)", ref)
    if not m:
        raise ValueError(f"Invalid cell ref: {ref}")
    return int(m.group(2)), col_to_num(m.group(1))


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if not np.isfinite(value):
            return None
        return float(value)
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, np.datetime64):
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def sector_name(value):
    try:
        if pd.isna(value):
            return None
        return ICB_MAP.get(int(float(value)), str(value))
    except Exception:
        return str(value) if value is not None else None


def source_series(df: pd.DataFrame, header: str) -> pd.Series:
    if header in df.columns:
        return df[header]
    alias = ALIASES.get(header)
    if alias and alias in df.columns:
        return df[alias]
    return pd.Series([None] * len(df), index=df.index)


def build_report_table(df: pd.DataFrame, weight_col: str, headers: list[str]) -> pd.DataFrame:
    out = pd.DataFrame(index=df.index)
    for header in headers:
        if header == "ISIN":
            out[header] = df["ISIN"].astype(str)
        elif header == "LIBELLE":
            out[header] = source_series(df, "Name").fillna(df["ISIN"]).astype(str)
        elif header in {"%ACTIF", "%ACTIF 100%"}:
            out[header] = pd.to_numeric(df[weight_col], errors="coerce").fillna(0.0)
        elif header == "ICB19 Supersector":
            out[header] = source_series(df, "Secto").map(sector_name)
        elif header == "Hors indice":
            out[header] = 0
        elif header in {"Beta", "Contrib TE", "Contrib Alpha"}:
            out[header] = None
        else:
            out[header] = source_series(df, header)
    return out.sort_values("%ACTIF", ascending=False).reset_index(drop=True)


def build_data_table(df: pd.DataFrame, headers: list[str], perf_live: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame(index=df.index)
    for header in headers:
        if header == "ISIN":
            out[header] = df["ISIN"].astype(str)
        elif header == "Name":
            out[header] = source_series(df, "Name").fillna(df["ISIN"]).astype(str)
        elif header == " Benchmark ICB Supersector ":
            out[header] = source_series(df, header).map(sector_name)
        elif header in {"Perf5D", "Perf1M", "Perf3M", "Perf6M"}:
            out[header] = df["Company SEDOL"].astype(str).map(perf_live[header])
        elif header.endswith("_LAST") or header.endswith("_DELTA"):
            out[header] = None
        else:
            out[header] = source_series(df, header)
    return out.reset_index(drop=True)


def compute_return_maps(returns: pd.DataFrame, sedols: list[str], end_date: pd.Timestamp):
    ret = returns.copy()
    if not isinstance(ret.index, pd.DatetimeIndex):
        ret.index = pd.to_datetime(ret.index, errors="coerce")
    ret = ret.sort_index()
    end = ret.index[ret.index < end_date].max()
    windows = {
        "Perf 1W": end - pd.offsets.BDay(5),
        "Perf 1M": end - pd.offsets.BDay(21),
        "Perf YTD": pd.Timestamp(end.year, 1, 1),
        "Perf 1Y": end - pd.offsets.BDay(252),
        "Perf5D": end - pd.offsets.BDay(5),
        "Perf1M": end - pd.offsets.BDay(21),
        "Perf3M": end - pd.offsets.BDay(63),
        "Perf6M": end - pd.offsets.BDay(126),
    }
    cols = [s for s in sedols if s in ret.columns]
    maps = {}
    for name, start in windows.items():
        vals = (1 + ret.loc[start:end, cols].fillna(0)).prod() - 1 if cols else pd.Series(dtype="float64")
        maps[name] = vals.to_dict()
    return maps, end


def build_inputs():
    sec = pd.read_parquet(SEC_LIST_PATH).copy()
    opt = pd.read_parquet(OPT_RESULT_PATH).copy()
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    returns = pd.read_parquet(RETURNS_PATH)

    opt["ISIN"] = opt["ISIN"].astype(str)
    opt["Company SEDOL"] = opt["Company SEDOL"].astype(str)
    opt["Wopt"] = pd.to_numeric(opt["Wopt"], errors="coerce").fillna(0.0)
    bench_col = "Weight in STOXX EUROPE 600"
    opt[bench_col] = pd.to_numeric(opt[bench_col], errors="coerce").fillna(0.0)
    opt[bench_col] = opt[bench_col] / opt[bench_col].sum()

    sec_isins = set(sec["ISIN"].astype(str))
    fund_src = opt[opt["ISIN"].isin(sec_isins)].copy()
    bench_src = opt[opt[bench_col] > 0].copy()

    sedols = opt["Company SEDOL"].dropna().astype(str).unique().tolist()
    perf_maps, perf_end = compute_return_maps(returns, sedols, pd.Timestamp("2025-06-01"))
    perf_live = pd.DataFrame(index=sedols)
    for col in ["Perf5D", "Perf1M", "Perf3M", "Perf6M"]:
        perf_live[col] = pd.Series(perf_maps[col])

    report_headers = template_headers("Fonds")
    data_headers = template_headers("DATA")
    fund = build_report_table(fund_src, "Wopt", report_headers)
    bench = build_report_table(bench_src, bench_col, report_headers)
    fund["Hors indice"] = (~fund["ISIN"].isin(set(bench["ISIN"]))).astype(int)
    data = build_data_table(opt, data_headers, perf_live)

    active = opt[["ISIN", "Name", "Company SEDOL", "Wopt", bench_col, "Score ML", "Secto", "Exchange Country Region"]].copy()
    active["Portfolio Weight"] = active["Wopt"]
    active["Benchmark Weight"] = active[bench_col]
    active["Active Weight"] = active["Portfolio Weight"] - active["Benchmark Weight"]
    active["ICB19 Supersector"] = active["Secto"].map(sector_name)
    active["In Portfolio"] = active["Portfolio Weight"] > 0.0001

    dev = active[["ISIN", "Name", "Portfolio Weight", "Benchmark Weight"]].copy()
    dev["Deviation"] = dev["Portfolio Weight"] - dev["Benchmark Weight"]
    dev = dev.sort_values("Deviation")
    top_under = dev.head(5)[["Name", "Deviation"]].rename(columns={"Name": "LIBELLE"})
    top_over = dev.tail(5)[["Name", "Deviation"]].rename(columns={"Name": "LIBELLE"})
    hors = fund[fund["Hors indice"] == 1][["ISIN", "LIBELLE", "%ACTIF"]].sort_values("%ACTIF", ascending=False)
    worst_ml = fund[pd.to_numeric(fund["Score ML"], errors="coerce").fillna(0) < 1][["LIBELLE", "Score ML"]].sort_values("Score ML")

    def perf_table(src: pd.DataFrame, weight_col: str, n: int, largest: bool):
        t = src[["ISIN", "Name", "Company SEDOL", weight_col, "Secto"]].copy()
        t["%ACTIF"] = t[weight_col]
        t["LIBELLE"] = t["Name"]
        t["ICB19 Supersector"] = t["Secto"].map(sector_name)
        for col in ["Perf 1W", "Perf 1M", "Perf YTD", "Perf 1Y"]:
            t[col] = t["Company SEDOL"].map(perf_maps[col])
        t = t[["ISIN", "%ACTIF", "LIBELLE", "ICB19 Supersector", "Perf 1W", "Perf 1M", "Perf YTD", "Perf 1Y"]]
        return (t.nlargest(n, "Perf 1W") if largest else t.nsmallest(n, "Perf 1W")).reset_index(drop=True)

    ret = returns.copy()
    if not isinstance(ret.index, pd.DatetimeIndex):
        ret.index = pd.to_datetime(ret.index, errors="coerce")
    ret = ret[(ret.index >= pd.Timestamp("2025-06-01") - pd.DateOffset(years=1)) & (ret.index < pd.Timestamp("2025-06-01"))]
    by_sedol = active.groupby("Company SEDOL")[["Portfolio Weight", "Benchmark Weight"]].sum()
    by_sedol["Active"] = by_sedol["Portfolio Weight"] - by_sedol["Benchmark Weight"]
    common = [s for s in by_sedol.index.astype(str) if s in ret.columns]
    te = float(ret[common].fillna(0).dot(by_sedol.loc[common, "Active"]).std(ddof=1) * math.sqrt(252))

    return {
        "fund": fund,
        "bench": bench,
        "data": data,
        "top10_fund": perf_table(fund_src, "Wopt", 10, True),
        "worst10_fund": perf_table(fund_src, "Wopt", 10, False),
        "top20_bench": perf_table(bench_src, bench_col, 20, True),
        "worst20_bench": perf_table(bench_src, bench_col, 20, False),
        "top_under": top_under,
        "top_over": top_over,
        "hors": hors,
        "worst_ml": worst_ml,
        "metrics": {
            "te": te,
            "summary": summary,
            "fund_score": float((active["Portfolio Weight"] * active["Score ML"]).sum()),
            "bench_score": float((active["Benchmark Weight"] * active["Score ML"]).sum()),
            "active_share": float(0.5 * active["Active Weight"].abs().sum()),
            "holdings": int(active["In Portfolio"].sum()),
            "bench_members": int((active["Benchmark Weight"] > 0).sum()),
            "perf_end": str(pd.Timestamp(perf_end).date()),
        },
    }


def get_sheet_paths(zf: ZipFile) -> dict[str, str]:
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root.findall(f"{{{NS_PACKAGE_REL}}}Relationship")}
    paths = {}
    for sheet in wb_root.findall(f"{{{NS_MAIN}}}sheets/{{{NS_MAIN}}}sheet"):
        name = sheet.attrib["name"]
        rid = sheet.attrib[f"{{{NS_REL}}}id"]
        target = rels[rid]
        path = "xl/" + target.lstrip("/") if not target.startswith("/xl/") else target.lstrip("/")
        paths[name] = path.replace("xl//", "xl/")
    return paths


def set_cell_value(c: ET.Element, value):
    for tag in ["v", "is", "f"]:
        child = c.find(f"{{{NS_MAIN}}}{tag}")
        if child is not None:
            c.remove(child)
    c.attrib.pop("t", None)
    value = clean_value(value)
    if value is None or value == "":
        return
    if isinstance(value, bool):
        c.set("t", "b")
        v = ET.SubElement(c, f"{{{NS_MAIN}}}v")
        v.text = "1" if value else "0"
    elif isinstance(value, (int, float)):
        v = ET.SubElement(c, f"{{{NS_MAIN}}}v")
        v.text = repr(value)
    else:
        c.set("t", "inlineStr")
        is_el = ET.SubElement(c, f"{{{NS_MAIN}}}is")
        t_el = ET.SubElement(is_el, f"{{{NS_MAIN}}}t")
        t_el.text = str(value)


def sheet_maps(sheet_data: ET.Element):
    rows = {}
    cells = {}
    for row in sheet_data.findall(f"{{{NS_MAIN}}}row"):
        r = int(row.attrib.get("r", "0"))
        rows[r] = row
        for c in row.findall(f"{{{NS_MAIN}}}c"):
            ref = c.attrib.get("r")
            if ref:
                cells[ref] = c
    return rows, cells


def get_or_create_row(sheet_data: ET.Element, rows: dict[int, ET.Element], r: int) -> ET.Element:
    if r in rows:
        return rows[r]
    row_el = ET.Element(f"{{{NS_MAIN}}}row", {"r": str(r)})
    inserted = False
    for idx, existing in enumerate(list(sheet_data)):
        if int(existing.attrib.get("r", "0")) > r:
            sheet_data.insert(idx, row_el)
            inserted = True
            break
    if not inserted:
        sheet_data.append(row_el)
    rows[r] = row_el
    return row_el


def get_or_create_cell(sheet_data, rows, cells, r, cidx, style=None):
    ref = cell_ref(r, cidx)
    if ref in cells:
        return cells[ref]
    row_el = get_or_create_row(sheet_data, rows, r)
    attrs = {"r": ref}
    if style is not None:
        attrs["s"] = str(style)
    cell_el = ET.Element(f"{{{NS_MAIN}}}c", attrs)
    inserted = False
    for idx, existing in enumerate(list(row_el)):
        m = re.match(r"([A-Z]+)", existing.attrib.get("r", ""))
        if m and col_to_num(m.group(1)) > cidx:
            row_el.insert(idx, cell_el)
            inserted = True
            break
    if not inserted:
        row_el.append(cell_el)
    cells[ref] = cell_el
    return cell_el


def style_lookup(cells, col, row_preference):
    for row in row_preference:
        c = cells.get(cell_ref(row, col))
        if c is not None and "s" in c.attrib:
            return c.attrib["s"]
    return None


def set_dimension(root, max_row, max_col):
    dim = root.find(f"{{{NS_MAIN}}}dimension")
    if dim is not None:
        dim.set("ref", f"A1:{cell_ref(max_row, max_col)}")


def write_table_preserve(xml_bytes: bytes, df: pd.DataFrame) -> bytes:
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS_MAIN}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{NS_MAIN}}}sheetData")
    rows, cells = sheet_maps(sheet_data)
    matrix = [list(df.columns)] + df.where(pd.notna(df), None).values.tolist()
    max_existing_row = max(rows.keys()) if rows else 1
    max_row = max(max_existing_row, len(matrix))
    max_col = len(matrix[0]) if matrix else 1
    for r in range(1, max_row + 1):
        for cidx in range(1, max_col + 1):
            style = style_lookup(cells, cidx, [r, 2, 1])
            cell = get_or_create_cell(sheet_data, rows, cells, r, cidx, style)
            value = matrix[r - 1][cidx - 1] if r <= len(matrix) else None
            set_cell_value(cell, value)
    set_dimension(root, max_row, max_col)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def df_values(df: pd.DataFrame, include_header=True):
    rows = []
    if include_header:
        rows.append(list(df.columns))
    rows.extend(df.where(pd.notna(df), None).values.tolist())
    return rows


def write_range(xml_bytes: bytes, start_ref: str, values, clear_rows=None, clear_cols=None) -> bytes:
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{NS_MAIN}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{NS_MAIN}}}sheetData")
    rows, cells = sheet_maps(sheet_data)
    start_r, start_c = split_cell(start_ref)
    matrix = values if isinstance(values, list) else [[values]]
    row_count = clear_rows or len(matrix)
    col_count = clear_cols or max(len(r) for r in matrix)
    for rr in range(row_count):
        for cc in range(col_count):
            r = start_r + rr
            cidx = start_c + cc
            style = style_lookup(cells, cidx, [r, 2, 1])
            cell = get_or_create_cell(sheet_data, rows, cells, r, cidx, style)
            value = matrix[rr][cc] if rr < len(matrix) and cc < len(matrix[rr]) else None
            set_cell_value(cell, value)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def update_workbook_calc(xml_bytes: bytes) -> bytes:
    root = ET.fromstring(xml_bytes)
    calc = root.find(f"{{{NS_MAIN}}}calcPr")
    if calc is None:
        calc = ET.SubElement(root, f"{{{NS_MAIN}}}calcPr")
    calc.set("calcMode", "auto")
    calc.set("fullCalcOnLoad", "1")
    calc.set("forceFullCalc", "1")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def build_workbook(data):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with ZipFile(TEMPLATE, "r") as zin:
        sheet_paths = get_sheet_paths(zin)
        mods = {
            sheet_paths["Fonds"]: write_table_preserve(zin.read(sheet_paths["Fonds"]), data["fund"]),
            sheet_paths["Benchmark"]: write_table_preserve(zin.read(sheet_paths["Benchmark"]), data["bench"]),
            sheet_paths["DATA"]: write_table_preserve(zin.read(sheet_paths["DATA"]), data["data"]),
            "xl/workbook.xml": update_workbook_calc(zin.read("xl/workbook.xml")),
        }
        top_xml = zin.read(sheet_paths["TopWorst Perf"])
        top_xml = write_range(top_xml, "C4", df_values(data["top10_fund"], True), clear_rows=11, clear_cols=8)
        top_xml = write_range(top_xml, "L4", df_values(data["worst10_fund"], True), clear_rows=11, clear_cols=8)
        top_xml = write_range(top_xml, "C20", df_values(data["top20_bench"], True), clear_rows=21, clear_cols=8)
        top_xml = write_range(top_xml, "L20", df_values(data["worst20_bench"], True), clear_rows=21, clear_cols=8)
        mods[sheet_paths["TopWorst Perf"]] = top_xml

        analyse_xml = zin.read(sheet_paths["Analyse"])
        analyse_xml = write_range(analyse_xml, "D1", [["Ptf"]])
        analyse_xml = write_range(analyse_xml, "E1", [["Benchmark"]])
        analyse_xml = write_range(analyse_xml, "R1", [["Benchmark"]])
        analyse_xml = write_range(analyse_xml, "Q2", [[data["metrics"]["te"]]])
        analyse_xml = write_range(analyse_xml, "H54", df_values(data["hors"][["LIBELLE", "%ACTIF"]], False), clear_rows=20, clear_cols=2)
        analyse_xml = write_range(analyse_xml, "J52", [[float(data["hors"]["%ACTIF"].sum()) if len(data["hors"]) else 0.0]])
        analyse_xml = write_range(analyse_xml, "L54", df_values(data["worst_ml"], False), clear_rows=20, clear_cols=2)
        analyse_xml = write_range(analyse_xml, "L45", df_values(data["top_over"], False), clear_rows=5, clear_cols=2)
        analyse_xml = write_range(analyse_xml, "M43", [[float(data["top_over"]["Deviation"].sum())]])
        analyse_xml = write_range(analyse_xml, "H45", df_values(data["top_under"], False), clear_rows=5, clear_cols=2)
        analyse_xml = write_range(analyse_xml, "I43", [[float(data["top_under"]["Deviation"].sum())]])
        mods[sheet_paths["Analyse"]] = analyse_xml

        with ZipFile(OUT_PATH, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, mods.get(item.filename, zin.read(item.filename)))
    return sheet_paths, mods.keys()


def main():
    data = build_inputs()
    sheet_paths, changed = build_workbook(data)
    print("OUTPUT", OUT_PATH)
    print("changed_parts", sorted(changed))
    print("sheet_paths", {k: sheet_paths[k] for k in ["Analyse", "Fonds", "Benchmark", "DATA", "TopWorst Perf"]})
    print("holdings", data["metrics"]["holdings"])
    print("bench_members", data["metrics"]["bench_members"])
    print("fund_score", round(data["metrics"]["fund_score"], 6))
    print("bench_score", round(data["metrics"]["bench_score"], 6))
    print("active_share", round(data["metrics"]["active_share"], 6))
    print("te", round(data["metrics"]["te"], 6))


if __name__ == "__main__":
    main()
