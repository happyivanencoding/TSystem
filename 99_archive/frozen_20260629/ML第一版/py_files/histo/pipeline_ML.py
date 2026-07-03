import pandas as pd
import numpy as np
import xgboost as xgb
from math import *
from operator import itemgetter
import datetime
import pickle
from multiprocessing import Pool
from dateutil import relativedelta
import os
import sys
 
from sklearn.preprocessing import LabelEncoder, PolynomialFeatures
from sklearn.metrics import accuracy_score, f1_score
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier, ExtraTreesRegressor, RandomForestRegressor
from sklearn.model_selection import RandomizedSearchCV
from sklearn.impute import KNNImputer
from sklearn.pipeline import make_pipeline
# from sklearn.utils import class_weight
# from imblearn.over_sampling import RandomOverSampler
# from imblearn.over_sampling import SMOTE
# from imblearn.over_sampling import ADASYN
# from imblearn.under_sampling import NearMiss
from scipy.stats import rankdata
from sklearn.metrics import mean_squared_error, r2_score
import shap
# import tensorflow as tf
# from keras.models import Sequential
# from keras.layers import Dense, Normalization
# from keras import regularizers
 
from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl import Workbook
# from openpyxl import load_workbook
from itertools import product
 
import copy

import py_files.config as cf

def load_pickle_files(screen_path, returns_path):

    screen_agg = pd.read_pickle(screen_path)
    df_returns = pd.read_pickle(returns_path)
    df_returns.set_index(pd.to_datetime(df_returns.index), inplace=True)
    return screen_agg, df_returns

# NEW
def screen_cleaning(screen_agg, univ): 
    """
    filtrer universe qu'on a choisi, >0, non Nan, pas de duplicates
    """
    screen_clean = copy.deepcopy(screen_agg)
    screen_clean = screen_clean[screen_clean['Weight in ' + univ]>0]
    screen_clean.rename(columns={'Weight in ' + univ: 'Weight in univ'}, inplace=True)
    mask = screen_clean.loc[(screen_clean['Weight in univ']>0)*(screen_clean['Company SEDOL'].notna()),'Company SEDOL'].unique()
    screen_clean = screen_clean.loc[screen_clean['Company SEDOL'].isin(mask)]
    screen_clean['Date'] = pd.to_datetime(screen_clean['Date'])
    screen_clean.reset_index(inplace=True)

    screen_clean = screen_clean[~screen_clean[['Company SEDOL','Date','Region','Sector ICB19']].isna().any(axis=1)]

    screen_clean.set_index(['Date', 'Company SEDOL'], inplace = True)
    screen_clean = screen_clean[~screen_clean.index.duplicated(keep='last')]
    screen_clean.reset_index(inplace=True)

    return screen_clean


# NEW
def compute_returns(screen_agg, period_to_predict, df_returns, sector_neutral='ICB19', type='return'):
    # Be sure sedol and date are index
    if ('Company SEDOL' in list(screen_agg.columns)) and ('Date' in list(screen_agg.columns)):
        data_clean = screen_agg.set_index(['Company SEDOL', 'Date'])
    else:
        data_clean = copy.deepcopy(screen_agg)

    # Create name of columns
    col_stock_returns = f'Stock {period_to_predict}M return'
    col_neutral_returns = f'Neutral {period_to_predict}M return'

    data_clean = data_clean[data_clean['Weight in univ']>0]

    # calculer rdt forward with différent window pour chaque date
    # X[1] DATE, X[0] Sedol
    data_clean[col_stock_returns] = list(data_clean.index.map(lambda x: ((1+df_returns[x[1]:x[1] + pd.DateOffset(days=int(period_to_predict)*30)][x[0]]).cumprod()-1).iloc[-1]))
    data_clean['smooth_cap'] = data_clean['Benchmark Market Value Millions in EUR']**(1/3)

    if (sector_neutral == 'ICB19') or (sector_neutral == 'ICB11'): # par secteur => secteur neutre
        # Calculer par secteur les returns 
        df_sectors = data_clean.reset_index(level=1).groupby(['Date', 'Sector ' + sector_neutral]).apply(lambda x: (x['Weight in univ'].dot(x[[col_stock_returns]]))/x['Weight in univ'].sum())
        df_sectors.columns= [col_neutral_returns]
        # Calculer dénominateur pour Contrib par secteur
        df_sectors[f'm_cap_bench {period_to_predict}M'] = data_clean.reset_index(level=1).groupby(['Date', 'Sector ' + sector_neutral])['smooth_cap'].sum()
        data_clean.reset_index(inplace=True)
        df_sectors.reset_index(inplace=True)
        data_clean = data_clean.merge(df_sectors,how='left', on=['Date', 'Sector ' + sector_neutral]).set_index(['Company SEDOL','Date'])
    else: # marché neutre 
        df_market = data_clean.reset_index(level=1).groupby(['Date']).apply(lambda x: (x['Weight in univ'].dot(x[[col_stock_returns]]))/x['Weight in univ'].sum())
        df_market.columns= [col_neutral_returns]
        df_market[f'm_cap_bench {period_to_predict}M'] = data_clean.reset_index(level=1).groupby(['Date'])['smooth_cap'].sum()
        data_clean.reset_index(inplace=True)
        df_market.reset_index(inplace=True)
        data_clean = data_clean.merge(df_market,how='left', on=['Date']).set_index(['Company SEDOL','Date'])

    # rdt relative
    data_clean['Relative ' +str(period_to_predict) + 'M return'] = data_clean['Stock ' +str(period_to_predict) + 'M return'] - data_clean['Neutral ' +str(period_to_predict) + 'M return']
    # Contrib sectoriel
    data_clean['contrib ' +str(period_to_predict) + 'M' ] = (data_clean['smooth_cap']/data_clean[f'm_cap_bench {period_to_predict}M'])*data_clean['Relative ' +str(period_to_predict) + 'M return']
    data_clean = data_clean.drop([f'm_cap_bench {period_to_predict}M', 'smooth_cap'], axis = 1)
    # info ratio => exp pour pénaliser fort
    
    # VOL CSTE
    # data_clean['information_ratio '+ str(period_to_predict) + 'M'] = data_clean['Relative ' +str(period_to_predict) + 'M return']/np.exp(data_clean['Relative ' +str(period_to_predict) + 'M return'].std())
    # vol passé
    
    # GOOD ONE
    rolling_std = data_clean.groupby(level="Company SEDOL")['Relative ' +str(period_to_predict) + 'M return'].expanding().std().droplevel(0)
    data_clean['information_ratio '+ str(period_to_predict) + 'M'] = data_clean['Relative ' +str(period_to_predict) + 'M return']/np.exp(rolling_std)
    # vol future
    
    return data_clean


def fill_nan_values(data_input, features, period_to_predict, min_pct_avail_features=0.6, method='mean', obs_weight = 'none', returns_type = 'contrib'):

    data_input['Dividend Avg Percentile'] = data_input['Dividend Avg Percentile'].fillna(0)
    min_avail_features = ceil(len(features)*min_pct_avail_features)
    data_input['keep'] = data_input[features].isna().sum(axis=1) <= len(features) - min_avail_features
    col_contrib = ['contrib ' +str(period)+ 'M' for period in period_to_predict]
    to_fill = features + col_contrib if obs_weight == 'contrib' or returns_type == 'contrib'  else features

    if method == 'mean':
        data_input[to_fill] = data_input[to_fill].fillna(data_input[to_fill].mean())
    elif method == 'KNN Imputer':
        imputer = KNNImputer() 
        data_input[to_fill] = imputer.fit_transform(data_input[to_fill])
    else:
        data_input[to_fill] = data_input[to_fill].fillna(data_input[to_fill].median())

    data_input = data_input[data_input['keep'] == True]
    data_input = data_input.drop(columns=['keep'])

    return data_input

def add_features_variation(df, features, time_periods):

    col_name_list=[]
    for feature in features:
        for period in time_periods:
            col_name = f"{feature}_change_{int(period)}M"
            col_name_list.append(col_name)
            df[col_name] = df.groupby(level=0)[feature].pct_change(periods=int(period))
            df[col_name].replace({np.inf: 1, -np.inf: -1}, inplace=True)
    return df, col_name_list
    
def labellize_data(df, period_to_predict, classes, returns_type, type_label):

    if type_label=='classification':
        if returns_type == 'returns':
            for date in df.index.get_level_values('Date').unique():
                df_date = df[df.index.get_level_values('Date') == date]
                pct_rank = df_date[f'Relative {period_to_predict}M return'].rank(pct=True,ascending=False)
                for class_ in list(classes.keys()):
                    class_index = pct_rank[(pct_rank>=classes[class_][0])*(pct_rank<=classes[class_][1])].index
                    df.loc[class_index, f'{period_to_predict}M label'] = class_
        elif returns_type == 'info_ratio':
            for date in df.index.get_level_values('Date').unique():
                df_date = df[df.index.get_level_values('Date') == date]
                pct_rank = df_date[f'information_ratio {period_to_predict}M'].rank(pct=True,ascending=False)
                for class_ in list(classes.keys()):
                    class_index = pct_rank[(pct_rank>=classes[class_][0])*(pct_rank<=classes[class_][1])].index
                    df.loc[class_index, f'{period_to_predict}M label'] = class_
        elif returns_type == 'contrib':
            for date in df.index.get_level_values('Date').unique():
                df_date = df[df.index.get_level_values('Date') == date]
                pct_rank = df_date[f'contrib {period_to_predict}M'].rank(pct=True,ascending=False)
                for class_ in list(classes.keys()):
                    class_index = pct_rank[(pct_rank>=classes[class_][0])*(pct_rank<=classes[class_][1])].index
                    df.loc[class_index, f'{period_to_predict}M label'] = class_
    elif type_label=='regression':
        if returns_type=='returns':
            for period in period_to_predict:
                df[f'{period}M label'] = df[f'Relative {period}M return']
        elif returns_type == 'info_ratio':
            for period in period_to_predict:
                df[f'{period}M label'] = df[f'information_ratio {period}M']
        elif returns_type == 'contrib':
            for period in period_to_predict:
                df[f'{period}M label'] = df[f'contrib {period}M']

    return df

def predict(df_scores, model_name, model_params, period_to_predict, include_columns, split_data_date, monotone_constraints, sampling = 'base', obs_weight = 'none', type_label = 'none'):

    training_set = df_scores.loc[df_scores.index.get_level_values('Date') < split_data_date]
    test_set = df_scores.loc[df_scores.index.get_level_values('Date') >= split_data_date]
   
    y_pred=[]
    for period in period_to_predict:
        training_set_na_null = training_set.copy(deep=True).dropna(subset=[f"{period}M label"])
        if type_label == 'classification':
            le = LabelEncoder()
            training_set[f"{period}M label"] = le.fit_transform(training_set[f"{period}M label"])
            test_set[f"{period}M label"] = le.transform(test_set[f"{period}M label"])
 
        X_train = training_set_na_null.drop(f"{period}M label", axis=1)
        y_train = training_set_na_null[f"{period}M label"]

        X_train = X_train[include_columns]
        X_test = test_set[include_columns]
        y_test = test_set[f"{period}M label"]
   
        model = None
 
        if model_name == 'XGBoostOptimizer':
            if type_label == 'classification':
                xgb_model = xgb.XGBClassifier()
            elif type_label == 'regression':
                xgb_model = xgb.XGBRegressor()
            for param in model_params:
                if type(model_params[param])!=list:
                    model_params[param] = [model_params[param]]
            model = RandomizedSearchCV(estimator=xgb_model, param_distributions=model_params, n_iter=1000, cv=5, scoring='neg_mean_squared_error', random_state=42, n_jobs=-1)
            #model.fit(X_train, y_train, sample_weight=classes_weights).best_params_
 
        elif model_name == 'XGBoost' or model_name == 'XGBoost1' or model_name == 'XGBoost2':
            if type_label=='classification':
                model = xgb.XGBClassifier(
                subsample=model_params['subsample'],
                reg_lambda=model_params['reg_lambda'],
                reg_alpha=model_params['reg_alpha'],
                random_state=42,
                n_estimators=model_params['n_estimators'],
                min_child_weight=model_params['min_child_weight'],
                max_depth=model_params['max_depth'],
                learning_rate=model_params['learning_rate'],
                gamma=model_params['gamma'],
                colsample_bytree=model_params['colsample_bytree'],
                monotone_constraints=monotone_constraints,
                objective = model_params['objective']
                #class_weight={0:4, 1:4, 2:1}
            )
            
            ############ Model à Utiliser #####################
            elif type_label=='regression':
                model = xgb.XGBRegressor(
                    subsample=model_params['subsample'],
                    reg_lambda=model_params['reg_lambda'],
                    reg_alpha=model_params['reg_alpha'],
                    random_state=42,
                    n_estimators=model_params['n_estimators'],
                    min_child_weight=model_params['min_child_weight'],
                    max_depth=model_params['max_depth'],
                    learning_rate=model_params['learning_rate'],
                    gamma=model_params['gamma'],
                    colsample_bytree=model_params['colsample_bytree'],
                    monotone_constraints=monotone_constraints,
                    #class_weight={0:4, 1:4, 2:1}
                    )
 
        elif model_name == 'Logistic Regression Optimizer':
            parameters = {
                'penalty': ['l2','none'],
                'C': [0.001, 0.01, 0.1, 1, 10, 100],
                'fit_intercept': [True, False],
                'max_iter': [50, 100, 200, 300, 500, 1000],
                'solver': ['newton-cg', 'lbfgs', 'sag', 'saga'],
                'random_state': [42],
                'class_weight': ['balanced']
            }
 
            lr_model = LogisticRegression()
            model = RandomizedSearchCV(estimator=lr_model, param_distributions=parameters, n_iter=50, cv=3, scoring='f1_weighted', random_state=42)
            print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train).best_params_)
 
        elif model_name == 'Logistic Regression':
            model = LogisticRegression(
                solver='liblinear',
                random_state=42,
                penalty='l2',
                max_iter=500,
                fit_intercept=True,
                class_weight='balanced',
                C=0.01
            )
 
        elif model_name == 'Random Forest Optimizer':
            parameters = {
                'n_estimators': [int(x) for x in range(100, 1001, 100)],
                'criterion': ['gini', 'entropy', 'log_loss'],
                'max_depth': [int(x) for x in range(2, 6)],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 4],
                'bootstrap': [False, True],
                'warm_start': [False, True],
                'random_state' : [42]
                #'class_weight': ['balanced']
            }
            rf_model = RandomForestClassifier()
            model = RandomizedSearchCV(estimator=rf_model, param_distributions=parameters, n_iter=50, cv=3, scoring='f1_weighted', random_state=42)
            print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train, sample_weight=classes_weights).best_params_)
 
        elif model_name == 'Random Forest':
            model = RandomForestClassifier(
                warm_start=False,
                random_state=42,
                n_estimators=300,
                # min_samples_split=10,
                # min_samples_leaf=1,
                max_depth=5,
                criterion='gini',
                #class_weight='balanced',
                bootstrap=True
            )
 
        elif model_name == 'Multilinear Regression Optimizer':
            poly = PolynomialFeatures(degree=2, include_bias=False)
            linear_model = LinearRegression()
            pipeline = make_pipeline(poly, linear_model)
 
            parameters = {
                'polynomialfeatures__degree': [1, 2, 3],
                'linearregression__fit_intercept': [True, False],  
            }
 
            model = RandomizedSearchCV(
                pipeline,
                param_distributions=parameters,
                n_iter=10,
                scoring='neg_mean_squared_error',
                cv=5,
                random_state=42
            )
            print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train).best_params_)
 
        elif model_name == 'Multilinear Regression':
            poly = PolynomialFeatures(degree=2, include_bias=False)
            linear_model = LinearRegression(fit_intercept=True)
            model = make_pipeline(poly, linear_model)
 
        elif model_name == 'SVC':
            model = SVC(
                class_weight='balanced',
                random_state=42
            )
 
        if type_label == 'classification':
            model.fit(X_train, y_train, sample_weight=classes_weights)
        elif type_label == 'regression':
            model.fit(X_train, y_train)
 
        if type_label == 'classification':
            if 'Optimizer' in model_name:
                shap_explainer = shap.TreeExplainer(model.best_estimator_)
            else:
                shap_explainer = shap.TreeExplainer(model)
            shap_values = shap_explainer.shap_values(X_test)
            class_names=list(le.inverse_transform(model.classes_))
            shap_val_df = pd.DataFrame()
            for i in range(len(class_names)):
                shap_val_df = pd.concat([shap_val_df, pd.DataFrame(data = np.abs(shap_values[i]).mean(0), index=include_columns, columns=[class_names[i]])], axis=1)
            shap_val_df['overall'] = shap_val_df.sum(axis=1)
            y_pred = model.predict(X_test)
            y_proba = model.predict_proba(X_test)
            f1 = f1_score(y_test, y_pred.round(), average='weighted')
            accuracy = accuracy_score(y_test, y_pred.round())
            if 'Regression' not in model_name:
                y_pred_labels = le.inverse_transform(y_pred)
       
        elif type_label == 'regression':
            if 'Optimizer' in model_name:
                shap_explainer = shap.TreeExplainer(model.best_estimator_)
            else:
                shap_explainer = shap.TreeExplainer(model)
            shap_values = shap_explainer.shap_values(X_test)
            shap_val_df = pd.DataFrame(data = np.abs(shap_values).mean(0), index=include_columns, columns=['overall'])
            new_pred = model.predict(X_test)
            y_pred.append(new_pred)
            # mse = mean_squared_error(y_test, new_pred)
            mse = 0
            # r2 = r2_score(y_test, new_pred)
            #r2_corr = 1 - (1- r2)*((len(y_test)-1)/(len(y_test)-len(include_columns)-1))
            r2_corr=0
 
    if 'Optimizer' in model_name:
        model_params = model.best_estimator_.get_params()
    else:
        model_params = model.get_params()
 
    model_params = list(itemgetter('subsample','reg_lambda','reg_alpha','random_state','n_estimators','min_child_weight','max_depth','learning_rate','gamma','colsample_bytree','monotone_constraints')(model_params))
    dict_params = dict(zip(['subsample','reg_lambda','reg_alpha','random_state','n_estimators','min_child_weight','max_depth','learning_rate','gamma','colsample_bytree','monotone_constraints'],model_params))
 
    if type_label == 'classification':
        return y_pred_labels,y_proba, shap_val_df, class_names, f1, accuracy, dict_params, [model,X_test]
    elif type_label == 'regression':
        return y_pred, mse

def dist_pred(y_true, y_pred):
 
    ranks_true = rankdata(y_true)
    ranks_true = len(ranks_true)-ranks_true+1
 
    ranks_pred = rankdata(y_pred)
    ranks_pred = len(ranks_pred)-ranks_pred+1
   
    dist = np.sum(np.abs(ranks_true-ranks_pred))
 
    return dist

def export_results_to_excel(wb, sheet_name, perf, buy_list_all, parameters, score1, score2, shap_values, type_label):

    wb.create_sheet(sheet_name)
    ws = wb.get_sheet_by_name(sheet_name)
    rows = dataframe_to_rows(perf, index=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    if type_label == 'classification':
        ws.cell(row=1,column=len(perf.columns)+3,value = 'f1 score')
        ws.cell(row=2,column=len(perf.columns)+3,value = 'accuracy')
    if type_label == 'regression':
        ws.cell(row=1,column=len(perf.columns)+3,value = 'MSE')
        ws.cell(row=2,column=len(perf.columns)+3,value = 'R2 corrigé')
    ws.cell(row=1,column=len(perf.columns)+4,value = score1)
    ws.cell(row=2,column=len(perf.columns)+4,value = score2)

    rows = dataframe_to_rows(shap_values, index=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=3+r_idx, column=len(perf.columns)+2+c_idx, value=value)
    
    ws.cell(row=7+len(shap_values),column=len(perf.columns)+3,value = 'parameters')
    for i,param in enumerate(parameters):
        ws.cell(row=7+len(shap_values)+i,column=len(perf.columns)+4,value = str(param))

    width_buy_list = len(buy_list_all[list(buy_list_all)[0]].columns)+1
    shift_col_buy_list = 0
    for buy_list in buy_list_all.keys():
        ws.cell(row=1, column=len(perf.columns)+len(shap_values.columns)+5+shift_col_buy_list, value=buy_list)
        rows = dataframe_to_rows(buy_list_all[buy_list], index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx+1, column=c_idx +len(perf.columns)+len(shap_values.columns)+4+shift_col_buy_list, value=value)
        shift_col_buy_list += width_buy_list

def launch_ML_process(strat_ML_params_file):
    #global_params
    df_features_path = cf.df_features_path

    strat_ML_params_file = read_pickle(strat_ML_params_file)

    screen_agg = pd.read_pickle(df_features_path)

    strat_params=strat_ML_params_file[cf.strat_ML]
    param_algo = cf.params[strat_params['algo']]
    classes={}
    for class_str in strat_params['classes']:
        class_ = class_str.split(';')
        classes[class_[0]] = np.array(class_[1].split('-'),dtype=float)
    monotone_constraints = dict(zip(strat_params['X'],strat_params['X_constraints']))
    for keys in monotone_constraints:
        monotone_constraints[keys] = int(monotone_constraints[keys])

    return launch_ML_prediction(screen_agg, strat_params['period_to_predict'],strat_params['Y'],strat_params['X'],
                        strat_params['min_avail_features'],strat_params['fill_na_X'],classes,strat_params['algo'],
                        param_algo,strat_params['sampling_method'],monotone_constraints,strat_params['training_window_yr'],
                        strat_params['test_window_mth'], strat_params['obs_weight'], strat_params['type_label'])

def launch_ML_prediction(screen_agg_origin, period_to_predict, returns_type, features, min_pct_avail_features, fill_na_method, classes, model_name,model_params, sampling_method, feature_constraints, training_window, test_window, obs_weight, type_label):

    screen_agg = copy.deepcopy(screen_agg_origin)

    screen_na_clean = fill_nan_values(screen_agg, features, period_to_predict, min_pct_avail_features, fill_na_method, obs_weight, returns_type)
   
    screen_label = labellize_data(screen_na_clean,period_to_predict,classes, returns_type, type_label)
    screen_label = screen_label[screen_label['Weight in univ']>0]
 
    last_date = screen_label.index.get_level_values('Date').max()
    cut_screen_date = last_date - relativedelta.relativedelta(years=training_window, months =test_window, day=1)
    screen_label = screen_label[screen_label.index.get_level_values('Date') >= cut_screen_date]

    first_date = screen_label.index.get_level_values('Date').min()
    split_date = first_date + relativedelta.relativedelta(years=training_window, months=1, day= 1)
    test_dataset = screen_label.loc[screen_label.index.get_level_values('Date') >= split_date].reset_index()
   
    y_pred, mse = predict(screen_label, model_name, model_params, period_to_predict, features, split_date, feature_constraints, sampling_method, obs_weight, type_label)

    for i, period in enumerate(period_to_predict):
                test_dataset['predicted_return_'+str(period)+'M'] = y_pred[i]

    col_predict=[]
    for period in period_to_predict:
        col_predict.append('predicted_return_'+str(period)+'M')
        test_dataset['predicted_return_'+str(period)+'M'] = test_dataset['predicted_return_'+str(period)+'M'].rank(ascending=True)

    test_dataset['Score ML'] = test_dataset[col_predict].mean(axis=1)
    test_dataset['Score ML'] = test_dataset.groupby('Date')['Score ML'].rank(pct=True, ascending=True)*10

    return test_dataset[test_dataset['Date'] == test_dataset['Date'].max()]

def format_ML_params(df_params, algo):

    param_algo = copy.deepcopy(df_params[algo])
    for param_name,param_val in param_algo.items():
        if ';' in str(param_val):
            param_str = param_val.split(';')
            param_algo[param_name] = [eval(p) for p in param_str]
    return param_algo.to_dict()

# def preprocess_data():
#     #global_params
#     screen_path = cf.screen_path
#     returns_path = cf.returns_path
#     df_features_path = cf.df_features_path
#     univ = cf.univ

    
#     preprocessing = cf.params_preprocessing
#     screen_agg, df_returns = load_pickle_files(screen_path, returns_path)

#     screen_agg.sort_values(by='Date', inplace=True)

#     features = preprocessing['X']
#     returns_horizon = preprocessing['returns_horizon']
#     variations_freq = preprocessing['variations_freq']
#     returns_type = preprocessing['Y']
#     returns_neutral = preprocessing['returns_neutral']

#     screen_agg.rename(columns={"Exchange Country Region":"Region",
#                             " Benchmark ICB Supersector ":"Sector ICB19",
#                             " Benchmark ICB Industry ":"Sector ICB11"},inplace=True)
#     screen_clean = screen_cleaning(screen_agg, univ)
#     for horizon in returns_horizon:
#         screen_clean = compute_returns(screen_clean, horizon, df_returns, returns_neutral, returns_type)
#     screen_full_feat, new_features = add_features_variation(screen_clean, features, variations_freq)
    
#     secto_dummies = pd.get_dummies(screen_full_feat['Sector ICB19'])
#     secto_dummies.columns=['Sector ' + str(i+1) for i in range(19)]
#     screen_full_feat = pd.concat([screen_full_feat,secto_dummies],axis=1)

#     screen_full_feat.to_pickle(df_features_path)

# def preprocess_data(preprocessing_pickle):
#     #global_params
#     screen_path = cf.screen_path
#     returns_path = cf.returns_path
#     df_features_path = cf.df_features_path
#     univ = cf.univ

#     preprocessing_list = read_pickle(preprocessing_pickle)
#     preprocessing = preprocessing_list[cf.preprocessing_id]

#     screen_agg, df_returns = load_pickle_files(screen_path, returns_path)
#     screen_agg.sort_values(by='Date', inplace=True)

#     features = preprocessing['X']
#     returns_horizon = preprocessing['returns_horizon']
#     variations_freq = preprocessing['variations_freq']
#     returns_type = preprocessing['Y']
#     returns_neutral = preprocessing['returns_neutral']

#     screen_agg.rename(columns={"Exchange Country Region":"Region",
#                             " Benchmark ICB Supersector ":"Sector ICB19",
#                             " Benchmark ICB Industry ":"Sector ICB11"},inplace=True)
#     screen_clean = screen_cleaning(screen_agg, univ)
#     for horizon in returns_horizon:
#         screen_clean = compute_returns(screen_clean, horizon, df_returns, returns_neutral, returns_type)
#     screen_full_feat, new_features = add_features_variation(screen_clean, features, variations_freq)

#     secto_dummies = pd.get_dummies(screen_full_feat['Sector ICB19'])
#     secto_dummies.columns=['Sector ' + str(i+1) for i in range(19)]
#     screen_full_feat = pd.concat([screen_full_feat,secto_dummies],axis=1)

#     screen_full_feat.to_pickle(df_features_path)    

def get_items_params_without_nan(df):
    lst = df.dropna().tolist()
    return lst[0] if len(lst) == 1 else lst

def preprocess_data(params_principal, params_preprocessing):
    #global_params
    screen_path = params_principal.loc['screen_path', 'param']
    returns_path = params_principal.loc['returns_path', 'param']
    df_features_path = params_principal.loc['df_features_path', 'param']
    univ = params_principal.loc['univ', 'param']

    features = get_items_params_without_nan(params_preprocessing['X'])
    returns_horizon = get_items_params_without_nan(params_preprocessing['returns_horizon'])
    variations_freq = get_items_params_without_nan(params_preprocessing['variations_freq'])
    returns_type = get_items_params_without_nan(params_preprocessing['Y']) 
    returns_neutral = get_items_params_without_nan(params_preprocessing['returns_neutral'])  

    screen_agg, df_returns = load_pickle_files(screen_path, returns_path)
    screen_agg.sort_values(by='Date', inplace=True)

    screen_agg.rename(columns={"Exchange Country Region":"Region",
                            " Benchmark ICB Supersector ":"Sector ICB19",
                            " Benchmark ICB Industry ":"Sector ICB11"},inplace=True)
    screen_clean = screen_cleaning(screen_agg, univ)
    for horizon in returns_horizon:
        screen_clean = compute_returns(screen_clean, horizon, df_returns, returns_neutral, returns_type)
    screen_full_feat, new_features = add_features_variation(screen_clean, features, variations_freq)

    secto_dummies = pd.get_dummies(screen_full_feat['Sector ICB19'])
    secto_dummies.columns=['Sector ' + str(i+1) for i in range(19)]
    screen_full_feat = pd.concat([screen_full_feat,secto_dummies],axis=1)

    screen_full_feat.to_pickle(df_features_path)

def save_mapping(file,strat_file_path,preprocessing_file_path):
    xl = pd.ExcelFile(file)
    strat_ML = xl.parse('Mapping_strat_ML', header=[0,1])
    preprocessing_df = xl.parse('Mapping_preprocessing', header=[0,1])

    unique_strat = list(strat_ML.columns.get_level_values(0).unique())
    dict_strat = {}
    for strat in unique_strat:
        dict_strat[strat] = {col: ((strat_ML[strat][col].dropna().values)[0] if len(list(strat_ML[strat][col].dropna().values))==1 else list(strat_ML[strat][col].dropna().values)) for col in strat_ML[strat].columns}

    unique_preprocessing = list(preprocessing_df.columns.get_level_values(0).unique())
    dict_preprocessing = {}
    for preprocessing in unique_preprocessing:
        dict_preprocessing[preprocessing] = {col: ((preprocessing_df[preprocessing][col].dropna().values)[0] if len(list(preprocessing_df[preprocessing][col].dropna().values))==1 else list(preprocessing_df[preprocessing][col].dropna().values)) for col in preprocessing_df[preprocessing].columns}

    file = open(strat_file_path, 'wb')
    pickle.dump(dict_strat, file)
    file.close()
    file = open(preprocessing_file_path, 'wb')
    pickle.dump(dict_preprocessing, file)
    file.close()

def read_pickle(file):

    """ Lit un fichier pickle """

    pkl_object = open(file, 'rb')
    object = pickle.load(pkl_object)
    pkl_object.close()
    return object