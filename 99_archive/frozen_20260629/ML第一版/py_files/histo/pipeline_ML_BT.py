import pandas as pd
import numpy as np
import xgboost as xgb
from math import *
from operator import itemgetter
import datetime
import pickle
from multiprocessing import Pool
from dateutil import relativedelta
import os
import sys
import copy

from sklearn.preprocessing import LabelEncoder, PolynomialFeatures
from sklearn.metrics import accuracy_score, f1_score
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier, ExtraTreesRegressor, RandomForestRegressor
from sklearn.model_selection import RandomizedSearchCV
from sklearn.impute import KNNImputer
from sklearn.pipeline import make_pipeline
from sklearn.utils import class_weight
from imblearn.over_sampling import RandomOverSampler
from imblearn.over_sampling import SMOTE
from imblearn.over_sampling import ADASYN
from imblearn.under_sampling import NearMiss
from scipy.stats import rankdata
from sklearn.metrics import mean_squared_error, r2_score
import shap
# import tensorflow as tf
# from keras.models import Sequential
# from keras.layers import Dense, Normalization
# from keras import regularizers
 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl import load_workbook
from itertools import product
 
# from inter_ML import *
 
# import sys
# sys.path.insert(0, r'\\groupe-ufg.com\commun\Prive\GestionAM\Ingenierie_Financiere\PROD\_EQUITY\Sec_List_Backtest')
from py_files.backtest_specific_ML import *

import py_files.config as cf


def load_pickle_files(screen_path, returns_path):

    screen_agg = pd.read_pickle(screen_path)
    df_returns = pd.read_pickle(returns_path)
    df_returns.set_index(pd.to_datetime(df_returns.index), inplace=True)
    return screen_agg, df_returns

def screen_cleaning(screen_agg, univ): 

    screen_clean = copy.deepcopy(screen_agg)
    # screen_clean = screen_clean[screen_clean['Weight in ' + univ]>0]
    screen_clean.rename(columns={'Weight in ' + univ: 'Weight in univ'}, inplace=True)
    mask = screen_clean.loc[(screen_clean['Weight in univ']>0)*(screen_clean['Company SEDOL'].notna()),'Company SEDOL'].unique()
    screen_clean = screen_clean.loc[screen_clean['Company SEDOL'].isin(mask)]
    screen_clean['Date'] = pd.to_datetime(screen_clean['Date'])
    screen_clean.reset_index(inplace=True)

    screen_clean = screen_clean[~screen_clean[['Company SEDOL','Date','Region','Sector ICB19']].isna().any(axis=1)]

    screen_clean.set_index(['Date', 'Company SEDOL'], inplace = True)
    screen_clean = screen_clean[~screen_clean.index.duplicated(keep='last')]
    screen_clean.reset_index(inplace=True)

    identif = screen_clean[['Company SEDOL', 'ISIN', 'Symbol', 'Name', 'Exchange Country Name', 'FactSet Ind', 'FactSet Economy', 'Curncy Iso', 'Region', 'Benchmark Country English', 'Sector ICB11', 'Sector ICB19', 'Benchmark Identifier -  SEDOLCHK', 'Exchange Country Iso2']].set_index('Company SEDOL')
    identif = identif[~identif.index.duplicated(keep='last')]
    identif.reset_index(inplace=True)
    list_sedol = screen_clean['Company SEDOL'].unique()
    list_date = screen_clean['Date'].unique()
    index = list(product(list_date, list_sedol))
    index = pd.DataFrame(index, columns = ['Date', 'Company SEDOL'])
    index = pd.merge(index, identif, how ='left', on = 'Company SEDOL')
    identif.drop(columns = 'Company SEDOL', inplace=True)
    screen_clean = pd.merge(index, screen_clean.drop(columns = identif.columns), how = 'left', on = ['Date', 'Company SEDOL'])
    screen_clean['Weight in univ'] = screen_clean['Weight in univ'].fillna(0)

    return screen_clean

def compute_returns(screen_agg, period_to_predict, df_returns, sector_neutral='ICB19', type='return'):

    if ('Company SEDOL' in list(screen_agg.columns)) and ('Date' in list(screen_agg.columns)):
        data_clean = screen_agg.set_index(['Company SEDOL', 'Date'])
    else:
        data_clean = copy.deepcopy(screen_agg)

    col_stock_returns = f'Stock {period_to_predict}M return'
    col_neutral_returns = f'Neutral {period_to_predict}M return'
    data_clean[col_stock_returns] = list(data_clean.index.map(lambda x: ((1+df_returns[x[1]:x[1] + pd.DateOffset(days=int(period_to_predict)*30)][x[0]]).cumprod()-1).iloc[-1]))
    data_clean['smooth_cap'] = data_clean['Benchmark Market Value Millions in EUR']**(1/3)

    if (sector_neutral == 'ICB19') or (sector_neutral == 'ICB11'):
        df_sectors = data_clean.reset_index(level=1).groupby(['Date', 'Sector ' + sector_neutral]).apply(lambda x: (x['Weight in univ'].dot(x[[col_stock_returns]]))/x['Weight in univ'].sum())
        df_sectors.columns= [col_neutral_returns]
        df_sectors[f'm_cap_bench {period_to_predict}M'] = data_clean.reset_index(level=1).groupby(['Date', 'Sector ' + sector_neutral])['smooth_cap'].sum()
        data_clean.reset_index(inplace=True)
        df_sectors.reset_index(inplace=True)
        data_clean = data_clean.merge(df_sectors,how='left', on=['Date', 'Sector ' + sector_neutral]).set_index(['Company SEDOL','Date'])
    else:
        df_market = data_clean.reset_index(level=1).groupby(['Date']).apply(lambda x: (x['Weight in univ'].dot(x[[col_stock_returns]]))/x['Weight in univ'].sum())
        df_market.columns= [col_neutral_returns]
        df_market[f'm_cap_bench {period_to_predict}M'] = data_clean.reset_index(level=1).groupby(['Date'])['smooth_cap'].sum()
        data_clean.reset_index(inplace=True)
        df_market.reset_index(inplace=True)
        data_clean = data_clean.merge(df_market,how='left', on=['Date']).set_index(['Company SEDOL','Date'])

    data_clean['Relative ' +str(period_to_predict) + 'M return'] = data_clean['Stock ' +str(period_to_predict) + 'M return'] - data_clean['Neutral ' +str(period_to_predict) + 'M return']
    data_clean['contrib ' +str(period_to_predict) + 'M' ] = (data_clean['smooth_cap']/data_clean[f'm_cap_bench {period_to_predict}M'])*data_clean['Relative ' +str(period_to_predict) + 'M return']
    data_clean = data_clean.drop([f'm_cap_bench {period_to_predict}M', 'smooth_cap'], axis = 1)
    data_clean['information_ratio '+ str(period_to_predict) + 'M'] = data_clean['Relative ' +str(period_to_predict) + 'M return']/np.exp(data_clean['Relative ' +str(period_to_predict) + 'M return'].std())
    
    return data_clean

def fill_nan_values(data_input, features, period_to_predict, min_pct_avail_features=0.6, method='mean', obs_weight = 'none', returns_type = 'contrib'):
    if 'Dividend Avg Percentile' in data_input.columns:
        data_input['Dividend Avg Percentile'] = data_input['Dividend Avg Percentile'].fillna(0)

    min_avail_features = ceil(len(features)*min_pct_avail_features)
    data_input['keep'] = data_input[features].isna().sum(axis=1) <= len(features) - min_avail_features
    col_contrib = ['contrib ' +str(period)+ 'M' for period in period_to_predict]
    to_fill = features + col_contrib if obs_weight == 'contrib' or returns_type == 'contrib'  else features

    if method == 'mean':
        data_input[to_fill] = data_input[to_fill].fillna(data_input[to_fill].mean())
    elif method == 'KNN Imputer':
        imputer = KNNImputer() 
        data_input[to_fill] = imputer.fit_transform(data_input[to_fill])
    else:
        data_input[to_fill] = data_input[to_fill].fillna(data_input[to_fill].median())

    data_input = data_input[data_input['keep'] == True]
    data_input = data_input.drop(columns=['keep'])

    return data_input

def add_features_variation(df, features, time_periods):

    col_name_list=[]
    for feature in features:
        for period in time_periods:
            col_name = f"{feature}_change_{int(period)}M"
            col_name_list.append(col_name)
            df[col_name] = df.groupby(level=0)[feature].pct_change(periods=int(period))
            df[col_name].replace({np.inf: 1, -np.inf: -1}, inplace=True)
    return df, col_name_list
    
def labellize_data(df, period_to_predict, classes, returns_type, type_label):

    df[f'{period_to_predict}M label'] = None

    if type_label=='classification':
        if returns_type == 'returns':
            for date in df.index.get_level_values('Date').unique():
                df_date = df[df.index.get_level_values('Date') == date]
                pct_rank = df_date[f'Relative {period_to_predict}M return'].rank(pct=True,ascending=False)
                for class_ in list(classes.keys()):
                    class_index = pct_rank[(pct_rank>=classes[class_][0])*(pct_rank<=classes[class_][1])].index
                    df.loc[class_index, f'{period_to_predict}M label'] = class_
        elif returns_type == 'info_ratio':
            for date in df.index.get_level_values('Date').unique():
                df_date = df[df.index.get_level_values('Date') == date]
                pct_rank = df_date[f'information_ratio {period_to_predict}M'].rank(pct=True,ascending=False)
                for class_ in list(classes.keys()):
                    class_index = pct_rank[(pct_rank>=classes[class_][0])*(pct_rank<=classes[class_][1])].index
                    df.loc[class_index, f'{period_to_predict}M label'] = class_
        elif returns_type == 'contrib':
            for date in df.index.get_level_values('Date').unique():
                df_date = df[df.index.get_level_values('Date') == date]
                pct_rank = df_date[f'contrib {period_to_predict}M'].rank(pct=True,ascending=False)
                for class_ in list(classes.keys()):
                    class_index = pct_rank[(pct_rank>=classes[class_][0])*(pct_rank<=classes[class_][1])].index
                    df.loc[class_index, f'{period_to_predict}M label'] = class_
    elif type_label=='regression':
        if returns_type=='returns':
            for period in period_to_predict:
                df[f'{period}M label'] = df[f'Relative {period}M return']
        elif returns_type == 'info_ratio':
            for period in period_to_predict:
                df[f'{period}M label'] = df[f'information_ratio {period}M']
        elif returns_type == 'contrib':
            for period in period_to_predict:
                df[f'{period}M label'] = df[f'contrib {period}M']

    return df

def predict(df_scores, 
            model_name, 
            model_params, 
            period_to_predict, 
            include_columns, 
            split_data_date, 
            monotone_constraints, 
            sampling = 'base', 
            obs_weight = 'none', 
            type_label = 'none'):

    # Split data into training and testing sets based on date
    # Rolling trainning and testing
    # if training_window_yr = 5, test_window_mth = 12, then model will use data of last 6 years, for the first 5 years as train set, for the last 12 months as test set
    training_set = df_scores.loc[df_scores.index.get_level_values('Date') < split_data_date]
    test_set = df_scores.loc[df_scores.index.get_level_values('Date') >= split_data_date]
    
    # training_set = training_set.dropna()
    y_pred=[] # This will contain predictions of returns in different months
    for period in period_to_predict: # for prod, period_to_predict : [3.0, 12.0]
        training_set_na_null = training_set.copy(deep=True).dropna(subset=[f"{period}M label"])

        if type_label == 'classification':
            le = LabelEncoder()
            training_set_na_null[f"{period}M label"] = le.fit_transform(training_set_na_null[f"{period}M label"])
            test_set[f"{period}M label"] = le.transform(test_set[f"{period}M label"])
 
        X_train = training_set_na_null.drop(f"{period}M label", axis=1)
        y_train = training_set_na_null[f"{period}M label"]
 
        if type_label == 'classification':
            if sampling == 'oversampling':
                oversample = RandomOverSampler(sampling_strategy='minority', random_state=42)
                for i in range(len(y_train.unique())-1):
                    X_train,y_train = oversample.fit_resample(X_train,y_train)
            elif sampling == 'undersampling':
                undersample = NearMiss(version=1)
                X_train,y_train = undersample.fit_resample(X_train,y_train)
            elif sampling == 'SMOTE':
                oversample = SMOTE(random_state=42)
                X_train,y_train = oversample.fit_resample(X_train,y_train)
            elif sampling == 'ADASYN':
                oversample = ADASYN(random_state=42)
                X_train,y_train = oversample.fit_resample(X_train,y_train)
 
        if type_label == 'classification':
            if obs_weight == 'contrib':
                classes_weights = np.array(np.abs(X_train[f'contrib {period}M']))
                classes_weights = classes_weights/classes_weights.sum()
            elif obs_weight == 'balanced':
                classes_weights = class_weight.compute_sample_weight(class_weight='balanced', y=y_train)
            else:
                classes_weights = np.ones(len(X_train))/len(X_train)

        X_train = X_train[include_columns]
        X_test = test_set[include_columns]
        y_test = test_set[f"{period}M label"]
   
        model = None  # Initiation of final model
 
        if model_name == 'XGBoostOptimizer':
            if type_label == 'classification':
                xgb_model = xgb.XGBClassifier()
            elif type_label == 'regression':
                xgb_model = xgb.XGBRegressor()
            for param in model_params:
                if type(model_params[param])!=list:
                    model_params[param] = [model_params[param]]
            model = RandomizedSearchCV(estimator=xgb_model, param_distributions=model_params, n_iter=1000, cv=5, scoring='neg_mean_squared_error', random_state=42, n_jobs=-1)
            #model.fit(X_train, y_train, sample_weight=classes_weights).best_params_
 
        elif model_name == 'XGBoost' or model_name == 'XGBoost1' or model_name == 'XGBoost2':
            if type_label=='classification':
                model = xgb.XGBClassifier(
                subsample=model_params['subsample'],
                reg_lambda=model_params['reg_lambda'],
                reg_alpha=model_params['reg_alpha'],
                random_state=42,
                n_estimators=model_params['n_estimators'],
                min_child_weight=model_params['min_child_weight'],
                max_depth=model_params['max_depth'],
                learning_rate=model_params['learning_rate'],
                gamma=model_params['gamma'],
                colsample_bytree=model_params['colsample_bytree'],
                monotone_constraints=monotone_constraints,
                objective = model_params['objective']
                #class_weight={0:4, 1:4, 2:1}
            )
               
            elif type_label=='regression':
                model = xgb.XGBRegressor(
                    subsample=model_params['subsample'],
                    reg_lambda=model_params['reg_lambda'],
                    reg_alpha=model_params['reg_alpha'],
                    random_state=42,
                    n_estimators=model_params['n_estimators'],
                    min_child_weight=model_params['min_child_weight'],
                    max_depth=model_params['max_depth'],
                    learning_rate=model_params['learning_rate'],
                    gamma=model_params['gamma'],
                    colsample_bytree=model_params['colsample_bytree'],
                    monotone_constraints=monotone_constraints,
                    #class_weight={0:4, 1:4, 2:1}
                    )
 
 
        elif model_name == 'Logistic Regression Optimizer':
            parameters = {
                'penalty': ['l2','none'],
                'C': [0.001, 0.01, 0.1, 1, 10, 100],
                'fit_intercept': [True, False],
                'max_iter': [50, 100, 200, 300, 500, 1000],
                'solver': ['newton-cg', 'lbfgs', 'sag', 'saga'],
                'random_state': [42],
                'class_weight': ['balanced']
            }
 
            lr_model = LogisticRegression()
            model = RandomizedSearchCV(estimator=lr_model, param_distributions=parameters, n_iter=50, cv=3, scoring='f1_weighted', random_state=42)
            print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train).best_params_)
 
        elif model_name == 'Logistic Regression':
            model = LogisticRegression(
                solver='liblinear',
                random_state=42,
                penalty='l2',
                max_iter=500,
                fit_intercept=True,
                class_weight='balanced',
                C=0.01
            )
 
        elif model_name == 'Random Forest Optimizer':
            parameters = {
                'n_estimators': [int(x) for x in range(100, 1001, 100)],
                'criterion': ['gini', 'entropy', 'log_loss'],
                'max_depth': [int(x) for x in range(2, 6)],
                'min_samples_split': [2, 5, 10],
                'min_samples_leaf': [1, 2, 4],
                'bootstrap': [False, True],
                'warm_start': [False, True],
                'random_state' : [42]
                #'class_weight': ['balanced']
            }
            rf_model = RandomForestClassifier()
            model = RandomizedSearchCV(estimator=rf_model, param_distributions=parameters, n_iter=50, cv=3, scoring='f1_weighted', random_state=42)
            print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train, sample_weight=classes_weights).best_params_)
 
        elif model_name == 'Random Forest':
            model = RandomForestClassifier(
                warm_start=False,
                random_state=42,
                n_estimators=300,
                # min_samples_split=10,
                # min_samples_leaf=1,
                max_depth=5,
                criterion='gini',
                #class_weight='balanced',
                bootstrap=True
            )
 
        elif model_name == 'Multilinear Regression Optimizer':
            poly = PolynomialFeatures(degree=2, include_bias=False)
            linear_model = LinearRegression()
            pipeline = make_pipeline(poly, linear_model)
 
            parameters = {
                'polynomialfeatures__degree': [1, 2, 3],
                'linearregression__fit_intercept': [True, False],  
            }
 
            model = RandomizedSearchCV(
                pipeline,
                param_distributions=parameters,
                n_iter=10,
                scoring='neg_mean_squared_error',
                cv=5,
                random_state=42
            )
            print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train).best_params_)
 
        elif model_name == 'Multilinear Regression':
            poly = PolynomialFeatures(degree=2, include_bias=False)
            linear_model = LinearRegression(fit_intercept=True)
            model = make_pipeline(poly, linear_model)
 
        elif model_name == 'SVC':
            model = SVC(
                class_weight='balanced',
                random_state=42
            )
        
        ### Model Fitting ###
        if type_label == 'classification':
            model.fit(X_train, y_train, sample_weight=classes_weights)
        elif type_label == 'regression':
            # print(f"{period}M label")
            # training_set_na_null.to_excel('training_set.xlsx')
            # X_train.to_excel('X_train.xlsx')
            # y_train.to_excel('y_train.xlsx')

            # training_set_na_null.to_pickle('training_set.pkl')
            # X_train.to_pickle('X_train.pkl')
            # y_train.to_pickle('y_train.pkl')

            model.fit(X_train, y_train)
        
        ### Model Prediction ###
        if type_label == 'classification':
            if 'Optimizer' in model_name:
                shap_explainer = shap.TreeExplainer(model.best_estimator_)
            else:
                shap_explainer = shap.TreeExplainer(model)
            shap_values = shap_explainer.shap_values(X_test)
            class_names=list(le.inverse_transform(model.classes_))
            shap_val_df = pd.DataFrame()
            for i in range(len(class_names)):
                shap_val_df = pd.concat([shap_val_df, pd.DataFrame(data = np.abs(shap_values[i]).mean(0), index=include_columns, columns=[class_names[i]])], axis=1)
            shap_val_df['overall'] = shap_val_df.sum(axis=1)

            y_pred = model.predict(X_test)
            y_proba = model.predict_proba(X_test)
            
            f1 = f1_score(y_test, y_pred.round(), average='weighted')
            accuracy = accuracy_score(y_test, y_pred.round())
            if 'Regression' not in model_name:
                y_pred_labels = le.inverse_transform(y_pred)
       
        elif type_label == 'regression':
            if 'Optimizer' in model_name:
                shap_explainer = shap.TreeExplainer(model.best_estimator_)
            else:
                shap_explainer = shap.TreeExplainer(model)
            shap_values = shap_explainer.shap_values(X_test)
            shap_val_df = pd.DataFrame(data = np.abs(shap_values).mean(0), index=include_columns, columns=['overall'])

            new_pred = model.predict(X_test)
            
            y_pred.append(new_pred)
            # mse = mean_squared_error(y_test, new_pred)
            mse = 0
            # r2 = r2_score(y_test, new_pred)
            r2 = 0
            #r2_corr = 1 - (1- r2)*((len(y_test)-1)/(len(y_test)-len(include_columns)-1))
            r2_corr=0
 
    if 'Optimizer' in model_name:
        model_params = model.best_estimator_.get_params()
    else:
        model_params = model.get_params()
 
    model_params = list(itemgetter('subsample','reg_lambda','reg_alpha','random_state','n_estimators','min_child_weight','max_depth','learning_rate','gamma','colsample_bytree','monotone_constraints')(model_params))
    dict_params = dict(zip(['subsample','reg_lambda','reg_alpha','random_state','n_estimators','min_child_weight','max_depth','learning_rate','gamma','colsample_bytree','monotone_constraints'],model_params))
 
    if type_label == 'classification':
        return y_pred_labels,y_proba, shap_val_df, class_names, f1, accuracy, dict_params, [model,X_test]
    elif type_label == 'regression':
        return y_pred, shap_val_df, mse, r2_corr, dict_params, [model,X_test]

def dist_pred(y_true, y_pred):
 
    ranks_true = rankdata(y_true)
    ranks_true = len(ranks_true)-ranks_true+1
 
    ranks_pred = rankdata(y_pred)
    ranks_pred = len(ranks_pred)-ranks_pred+1
   
    dist = np.sum(np.abs(ranks_true-ranks_pred))
 
    return dist

def export_results_to_excel(wb, sheet_name, perf, buy_list_all, parameters, score1, score2, shap_values, type_label):

    wb.create_sheet(sheet_name)
    ws = wb.get_sheet_by_name(sheet_name)
    rows = dataframe_to_rows(perf, index=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    if type_label == 'classification':
        ws.cell(row=1,column=len(perf.columns)+3,value = 'f1 score')
        ws.cell(row=2,column=len(perf.columns)+3,value = 'accuracy')
    if type_label == 'regression':
        ws.cell(row=1,column=len(perf.columns)+3,value = 'MSE')
        ws.cell(row=2,column=len(perf.columns)+3,value = 'R2 corrigé')
    ws.cell(row=1,column=len(perf.columns)+4,value = score1)
    ws.cell(row=2,column=len(perf.columns)+4,value = score2)

    rows = dataframe_to_rows(shap_values, index=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=3+r_idx, column=len(perf.columns)+2+c_idx, value=value)
    
    ws.cell(row=7+len(shap_values),column=len(perf.columns)+3,value = 'parameters')
    for i,param in enumerate(parameters):
        ws.cell(row=7+len(shap_values)+i,column=len(perf.columns)+4,value = str(param))

    width_buy_list = len(buy_list_all[list(buy_list_all)[0]].columns)+1
    shift_col_buy_list = 0
    for buy_list in buy_list_all.keys():
        ws.cell(row=1, column=len(perf.columns)+len(shap_values.columns)+5+shift_col_buy_list, value=buy_list)
        rows = dataframe_to_rows(buy_list_all[buy_list], index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx+1, column=c_idx +len(perf.columns)+len(shap_values.columns)+4+shift_col_buy_list, value=value)
        shift_col_buy_list += width_buy_list

def launch_ML_prediction(
        screen_agg_origin, 
        df_returns_origin, 
        univ, 
        period_to_predict, 
        returns_neutral, 
        returns_type, 
        features, 
        min_pct_avail_features, 
        fill_na_method, 
        create_feat_chg, 
        variations_frequency, 
        classes, 
        model_name,
        model_params, 
        sampling_method, 
        split_data_date, 
        feature_constraints, 
        class_to_backtest, 
        sector_neutral_bt, 
        ponderation_bt, 
        preprocessed_data, 
        rolling, 
        training_window, 
        test_window, 
        training_neutral, 
        training_filter, 
        obs_weight, 
        type_label, 
        allow_multiprocessing, 
        update_score_ML):

    screen_agg = copy.deepcopy(screen_agg_origin)
    df_returns = copy.deepcopy(df_returns_origin)
 
    if type(period_to_predict)!=list:
        period_to_predict= [period_to_predict]
    if preprocessed_data:
        screen_na_clean = fill_nan_values(screen_agg, features, period_to_predict, min_pct_avail_features, fill_na_method, obs_weight, returns_type)
    else:
        screen_clean = screen_cleaning(screen_agg, univ)
        screen_clean = compute_returns(screen_clean, period_to_predict, df_returns, returns_neutral, returns_type)
        screen_na_clean = fill_nan_values(screen_clean, features, period_to_predict, min_pct_avail_features, fill_na_method, returns_type)
 
        if create_feat_chg:
            screen_na_clean, new_features = add_features_variation(screen_na_clean, features, variations_frequency)
            features = features + new_features
   
    screen_label = labellize_data(screen_na_clean,period_to_predict,classes, returns_type, type_label)
    screen_label = screen_label[screen_label['Weight in univ']>0]
 
    if rolling:
        last_date = screen_label.index.get_level_values('Date').max()
        first_date = screen_label.index.get_level_values('Date').min()
        training_date = first_date + relativedelta.relativedelta(years=training_window)
        test_date = training_date + relativedelta.relativedelta(months=test_window)
        test_dataset = screen_label.loc[screen_label.index.get_level_values('Date') >= training_date].reset_index()
        screen_rolling = screen_label[(screen_label.index.get_level_values('Date')>=first_date)*(screen_label.index.get_level_values('Date')<=test_date)]
        list_params_rolling = [[screen_rolling, model_name, model_params, period_to_predict, features, training_date,feature_constraints, sampling_method, obs_weight, type_label]]
        while training_date + relativedelta.relativedelta(months=test_window)< last_date:
            first_date = first_date + relativedelta.relativedelta(months=test_window)
            training_date = training_date + relativedelta.relativedelta(months=test_window)
            test_date = training_date + relativedelta.relativedelta(months=test_window)
            screen_rolling = screen_label[(screen_label.index.get_level_values('Date')>=first_date)*(screen_label.index.get_level_values('Date')<=test_date)]            
            list_params_rolling.append([screen_rolling, model_name, model_params, period_to_predict, features, training_date, feature_constraints, sampling_method, obs_weight, type_label])
        if allow_multiprocessing:
            with Pool(os.cpu_count()-1) as p:
                results_period = p.starmap(predict, [params for params in list_params_rolling])
        else: 
            results_period = []
            for params in list_params_rolling:
                predict_result = predict(params[0],params[1],params[2],params[3],params[4],params[5],params[6],params[7], params[8], params[9])
                results_period.append(predict_result)
        if type_label =='classification':
            f1_periods = []
            accuracy_periods = []
            shap_val_periods = []
            for i, result in enumerate(results_period):
                train_date = list_params_rolling[i][5]
                test_date = train_date + relativedelta.relativedelta(months=test_window)
                test_dataset.loc[(test_dataset['Date'] >= train_date)*(test_dataset['Date'] <= test_date), 'predicted_class'] = result[0]
                test_dataset.loc[(test_dataset['Date'] >= train_date)*(test_dataset['Date'] <= test_date), result[3]] = result[1]
                f1_periods.append(result[4])
                accuracy_periods.append(result[5])
                shap_val_periods.append(result[2])
                inter_param = result[7]
            f1 = np.mean(f1_periods)
            accuracy = np.mean(accuracy_periods)
            parameters = results_period[-1][6]
            shap_val_df = shap_val_periods[-1]
        elif type_label == 'regression':
            mse_periods = []
            r2_corr_periods = []
            shap_val_periods = []
            for i, result in enumerate(results_period):
                train_date = list_params_rolling[i][5]
                test_date = train_date + relativedelta.relativedelta(months=test_window)
                for i, period in enumerate(period_to_predict):
                    test_dataset.loc[(test_dataset['Date'] >= train_date)*(test_dataset['Date'] <= test_date), 'predicted_return_'+str(period)+'M'] = result[0][i]
                mse_periods.append(result[2])
                r2_corr_periods.append(result[3])
                shap_val_periods.append(result[1])
                inter_param = result[5]
            mse = np.mean(mse_periods)
            r2_corr = np.mean(r2_corr_periods)
            parameters = results_period[-1][4]
            shap_val_df = shap_val_periods[-1]
 
    else:
        if type_label == 'classification':
            y_pred_labels, y_pred_proba, shap_val_df, classes_order, f1, accuracy, parameters, inter_param = predict(screen_label, model_name, model_params, period_to_predict, features, split_data_date,feature_constraints, sampling_method, obs_weight, type_label)
            test_dataset = screen_label.loc[screen_label.index.get_level_values('Date') >= split_data_date].reset_index()
            test_dataset['predicted_class'] = y_pred_labels
            test_dataset[classes_order] = y_pred_proba
        elif type_label == 'regression':
            y_pred, shap_val_df, mse, r2_corr, parameters, inter_param = predict(screen_label, model_name, model_params, period_to_predict, features, split_data_date,feature_constraints, sampling_method, obs_weight, type_label)
            test_dataset = screen_label.loc[screen_label.index.get_level_values('Date') >= split_data_date].reset_index()
            for i, period in enumerate(period_to_predict):
                test_dataset['predicted_return_'+str(period)+'M'] = y_pred[i]
   
    col_predict=[]
    for period in period_to_predict:
        col_predict.append('predicted_return_'+str(period)+'M')
        test_dataset['predicted_return_'+str(period)+'M'] = test_dataset['predicted_return_'+str(period)+'M'].rank(ascending=True)
        #test_dataset['predicted_return_'+str(period)+'M'] = (test_dataset['predicted_return_'+str(period)+'M']-test_dataset['predicted_return_'+str(period)+'M'].mean())/test_dataset['predicted_return_'+str(period)+'M'].std()
    test_dataset['predicted_return'] = test_dataset[col_predict].mean(axis=1)
    test_dataset['predicted_return'] = test_dataset.groupby('Date')['predicted_return'].rank(pct=True, ascending=True)*10
    # test_dataset.to_pickle('predictions.pkl')
   
    parameters['Model'] = model_name
    parameters['Rolling'] = rolling
    if rolling:
        parameters['Training window yr'] = training_window
        parameters['Test window mth'] = test_window
    parameters['Training neutral'] = training_neutral
    if training_neutral:
        parameters['Training filter'] = training_filter
    parameters['Target return'] = returns_type
    parameters['Training return horizon'] = period_to_predict
    parameters['Returns neutral'] = returns_neutral
    parameters['Univers'] = univ
    parameters['classes'] = str(classes)
    parameters['BT params'] = [sector_neutral_bt, ponderation_bt]
 
    bench = test_dataset[['ISIN', 'Date']]
    sec_list_dico={'bench': [bench, test_dataset]}
 
    if type_label == 'classification':
        for class_ in class_to_backtest:
            pct_class=classes[class_][1] - classes[class_][0]
            #sec_list by proba
            sec_list_proba = test_dataset.groupby('Date', group_keys=False).apply(lambda x: x.nlargest(ceil(pct_class*len(x)),class_))[['ISIN', 'Date']]
            sec_list_dico[class_ +' proba']=[sec_list_proba, test_dataset]
            #sec_list by label
            if len(test_dataset.loc[test_dataset["predicted_class"]==class_])>0:
                sec_list_label = test_dataset.loc[test_dataset["predicted_class"]==class_, ['ISIN', 'Date']]
            else :
                raise ValueError(f'The model did not predict any {class_}, for at least one period or sector.')
            sec_list_dico[class_ +' label']=[sec_list_label, test_dataset]
    elif type_label == 'regression':
        for date in test_dataset['Date'].unique():
            df_date = test_dataset[test_dataset['Date'] == date]
            test_dataset.loc[test_dataset['Date'] == date,'rank'] = df_date['predicted_return'].rank(pct=True,ascending=False)
        for class_ in class_to_backtest:
            min_rank = classes[class_][0]
            max_rank =  classes[class_][1]
            sec_list_pred = test_dataset.loc[(test_dataset['rank']<max_rank) & (test_dataset['rank']>min_rank), ['ISIN', 'Date']]
            sec_list_dico[class_ +' proba']=[sec_list_pred, test_dataset]  
    if update_score_ML:
        test_dataset.to_excel("SCORE_ML_2010_to_Today.xlsx")
        test_dataset.to_pickle("SCORE_ML_2010_to_Today.pkl")
    if type_label == 'classification':
        return [sec_list_dico,parameters,f1,accuracy,shap_val_df, inter_param]
    elif type_label == 'regression':
        return [sec_list_dico,parameters,mse,r2_corr,shap_val_df, inter_param]

def format_ML_params(df_params, algo):

    param_algo = copy.deepcopy(df_params[algo])
    for param_name,param_val in param_algo.items():
        if ';' in str(param_val):
            param_str = param_val.split(';')
            param_algo[param_name] = [eval(p) for p in param_str]
    return param_algo.to_dict()

def prepare_data_for_xl(result, df_returns, list_params):

    sec_list_names=[]
    list_params_bt = []
    
    for sec_list in result[0]:
        if sec_list == 'bench':
            list_params_bt.append([result[0][sec_list][0], "univ",result[0][sec_list][1],df_returns, 'Sector ICB19', 'Company SEDOL', 'ISIN', 'Date','Benchmark Market Value Millions in EUR',list_params[18],'mkt_cap'])
        else:
            list_params_bt.append([result[0][sec_list][0], "univ",result[0][sec_list][1],df_returns, 'Sector ICB19', 'Company SEDOL', 'ISIN', 'Date','Benchmark Market Value Millions in EUR',list_params[18],list_params[19]])
        sec_list_names.append(sec_list)
        
    with Pool(os.cpu_count()-1) as p:
        result_bt = p.starmap(backtest, [params for params in list_params_bt])

    col_bt = ['Perf '+sec_list for sec_list in sec_list_names]
    sec_list_bt_names = ['sec list '+sec_list for sec_list in sec_list_names]
    perf_list = [bt[0] for bt in result_bt]
    buy_list = [bt[1] for bt in result_bt]
    perf_all=pd.concat(perf_list,axis=1)
    perf_all.columns = col_bt
    buy_list_dict = dict(zip(sec_list_bt_names,buy_list))
    return perf_all, buy_list_dict

def launch_ML_process(strat_ML_params_file, output_file, features=None, update_score_ML=False):

    variables_inter = ['Value Avg Percentile','Quality Avg Percentile','Growth Avg Percentile']

    #global_params
    univ = cf.univ
    screen_path = cf.screen_path
    returns_path = cf.returns_path
    df_features_path = cf.df_features_path
    use_preprocessed_data = cf.use_preprocessed_data
    meta_model = cf.meta_model
    strat_ML_list = cf.strat_ML_list

    strat_ML_params_file = read_pickle(strat_ML_params_file)
    if use_preprocessed_data:
        screen_agg, df_returns = load_pickle_files(df_features_path, returns_path)
    else:
        screen_agg, df_returns = load_pickle_files(screen_path, returns_path)
        screen_agg.sort_values(by='Date', inplace=True)
        screen_agg.rename(columns={"Exchange Country Region":"Region",
                            " Benchmark ICB Supersector ":"Sector ICB19",
                            " Benchmark ICB Industry ":"Sector ICB11"},inplace=True)
        
    # output_file = "histo\output_ML_BT_" + datetime.datetime.now().strftime("%Y%m%d") + ".xlsx"
    wb = Workbook()

    list_params = []
    strat_names=[]
    for strat_ML in strat_ML_list:
        strat_params=strat_ML_params_file[strat_ML]
        param_algo = cf.params[strat_params['algo']]
        classes={}
        for class_str in strat_params['classes']:
            class_ = class_str.split(';')
            classes[class_[0]] = np.array(class_[1].split('-'),dtype=float)

        def create_mask_list(original_list):
            return [1] * (len(original_list) - 19) + [0] * 19
        if features != None:
            features = features
            X_constraints = create_mask_list(features)
            monotone_constraints = dict(zip(features,X_constraints))
        elif features == None:
            features = strat_params['X']
            monotone_constraints = dict(zip(features,strat_params['X_constraints']))

        for keys in monotone_constraints:
            monotone_constraints[keys] = int(monotone_constraints[keys])

        if strat_params['training_neutral']:
            list_params_neutral=[]
            unique_filter_val = screen_agg[strat_params['training_filter']].unique()
            for filter in unique_filter_val:
                screen_filtered = screen_agg[screen_agg[strat_params['training_filter']] == filter]
                list_params_neutral.append([screen_filtered, df_returns, univ, strat_params['period_to_predict'], strat_params['returns_neutral'], strat_params['Y'],features,
                                    strat_params['min_avail_features'],strat_params['fill_na_X'], strat_params['create_feature_change'],strat_params['variations_freq'],
                                    classes,strat_params['algo'],param_algo,strat_params['sampling_method'],strat_params['split_date'],monotone_constraints, 
                                    strat_params['class_to_backtest'], strat_params['sector_neutral_bt'],strat_params['ponderation_bt'], use_preprocessed_data,
                                    strat_params['rolling'],strat_params['training_window_yr'],strat_params['test_window_mth'], strat_params['training_neutral'], strat_params['training_filter'], strat_params['obs_weight'], strat_params['type_label'], 
                                    True, update_score_ML])
            with Pool(os.cpu_count()-1) as p:
                sub_result = p.starmap(launch_ML_prediction, [params for params in list_params_neutral])
                

            parameters=[]
            score1 = []
            score2 = []
            shap_val=[]
            list_sec_list = {}
            for i,result in enumerate(sub_result):
                parameters.append(result[1])
                score1.append(result[2])
                score2.append(result[3])
                result[4].index = result[4].index.map(lambda x: x+'_'+str(unique_filter_val[i]))
                shap_val.append(result[4])
                inter_param = result[5]
                for sec_list in result[0]:
                    if sec_list in list_sec_list.keys():
                        new_sec_list = pd.concat([list_sec_list[sec_list][0], result[0][sec_list][0]], axis=0, ignore_index=True)
                        new_bench = pd.concat([list_sec_list[sec_list][1], result[0][sec_list][1]], axis=0, ignore_index=True)
                        list_sec_list[sec_list] = [new_sec_list,new_bench]
                    else:
                        list_sec_list[sec_list] = [result[0][sec_list][0],result[0][sec_list][1]]

            full_result = [list_sec_list,parameters,np.mean(score1),np.mean(score2),pd.concat(shap_val)]
            perf_all, buy_list_dict = prepare_data_for_xl(full_result, df_returns, list_params_neutral[i])
            export_results_to_excel(wb,strat_ML, perf_all,buy_list_dict,full_result[1],full_result[2],full_result[3],full_result[4], strat_params['type_label'])
            # shap_values_graphs(inter_param[0], inter_param[1], wb, strat_ML, 'Y1')
            # for k in range(len(variables_inter)):
            #     print(variables_inter[k])
            #     marginal_effects(inter_param[0], inter_param[1], [variables_inter[k]], wb, strat_names[i], 'Y1' )
            #     for v in range(k+1, len(variables_inter)):
            #         marginal_effects(inter_param[0], inter_param[1], [variables_inter[k], variables_inter[v]], wb, strat_names[i], 'Y1' )
        else:
            list_params.append([screen_agg, df_returns, univ, strat_params['period_to_predict'], strat_params['returns_neutral'], strat_params['Y'], features,
                                strat_params['min_avail_features'],strat_params['fill_na_X'], strat_params['create_feature_change'],strat_params['variations_freq'],
                                classes,strat_params['algo'],param_algo,strat_params['sampling_method'],strat_params['split_date'],monotone_constraints, 
                                strat_params['class_to_backtest'], strat_params['sector_neutral_bt'],strat_params['ponderation_bt'], use_preprocessed_data,
                                strat_params['rolling'],strat_params['training_window_yr'],strat_params['test_window_mth'], strat_params['training_neutral'], strat_params['training_filter'], strat_params['obs_weight'], strat_params['type_label'], True])
            strat_names.append(strat_ML)        
        # launch_ML_prediction(screen_agg, df_returns, univ, strat_params['period_to_predict'], strat_params['returns_neutral'], strat_params['Y'],strat_params['X'],
        #                      strat_params['min_avail_features'],strat_params['fill_na_X'], strat_params['create_feature_change'],strat_params['variations_freq'],
        #                      classes,strat_params['algo'],param_algo,strat_params['sampling_method'],strat_params['split_date'],monotone_constraints, 
        #                      strat_params['class_to_backtest'], strat_params['sector_neutral_bt'],strat_params['ponderation_bt'], use_preprocessed_data)
    if len(list_params)>0:
        if len(list_params)==1:
            unique_result = launch_ML_prediction(list_params[0][0],list_params[0][1],list_params[0][2],list_params[0][3],list_params[0][4],list_params[0][5],
                                                 list_params[0][6],list_params[0][7],list_params[0][8],list_params[0][9],list_params[0][10],list_params[0][11],
                                                 list_params[0][12],list_params[0][13],list_params[0][14],list_params[0][15],list_params[0][16],list_params[0][17],
                                                 list_params[0][18],list_params[0][19],list_params[0][20],list_params[0][21],list_params[0][22],list_params[0][23],
                                                 list_params[0][24],list_params[0][25], list_params[0][26], list_params[0][27], True, update_score_ML)
            result_list = [unique_result]

        else:
            with Pool(os.cpu_count()-1) as p:
                result_list = p.starmap(launch_ML_prediction, [params for params in list_params])
        for i,result in enumerate(result_list):
            inter_param = result[-1]
            perf_all, buy_list_dict = prepare_data_for_xl(result, df_returns, list_params[i])
            export_results_to_excel(wb,strat_names[i], perf_all,buy_list_dict,[result[1]],result[2],result[3],result[4], strat_params['type_label'])
            # shap_values_graphs(inter_param[0], inter_param[1], wb, strat_names[i], 'Y1')
            # for k in range(len(variables_inter)):
            #     marginal_effects(inter_param[0], inter_param[1], [variables_inter[k]], wb, strat_names[i], 'Y1' )
            #     for v in range(k+1, len(variables_inter)):
            #         marginal_effects(inter_param[0], inter_param[1], [variables_inter[k], variables_inter[v]], wb, strat_names[i], 'Y1' )

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save(output_file)
    wb.close()
    path = '//groupe-ufg.com/commun/Prive/GestionAM/Ingenierie_Financiere/DOSSIERS_UTILISATEURS/BR/EQUITY/Louis/trash/'
    list_files = os.listdir(path)
    for file in list_files :
        os.remove(path + file)

def preprocess_data(param_file, preprocessing_pickle):

    xl = pd.ExcelFile(param_file)
    global_params = xl.parse('Launcher', nrows=7, header= None,index_col=0)
    process_params = xl.parse('Launcher', skiprows=8, usecols='C')

    #global_params
    screen_path = global_params.loc['screen_path'].iloc[0]
    returns_path = global_params.loc['returns_path'].iloc[0]
    df_features_path = global_params.loc['df_features_path'].iloc[0]
    univ = global_params.loc['univ'].iloc[0]
    preprocessing_id = process_params['Preprocessing'].iloc[0]
    preprocessing_list = read_pickle(preprocessing_pickle)
    preprocessing = preprocessing_list[preprocessing_id]
    screen_agg, df_returns = load_pickle_files(screen_path, returns_path)
    screen_agg.sort_values(by='Date', inplace=True)


    features = preprocessing['X']
    returns_horizon = preprocessing['returns_horizon']
    variations_freq = preprocessing['variations_freq']
    returns_type = preprocessing['Y']
    returns_neutral = preprocessing['returns_neutral']

    screen_agg.rename(columns={"Exchange Country Region":"Region",
                            " Benchmark ICB Supersector ":"Sector ICB19",
                            " Benchmark ICB Industry ":"Sector ICB11"},inplace=True)
    screen_clean = screen_cleaning(screen_agg, univ)
    for horizon in returns_horizon:
        screen_clean = compute_returns(screen_clean, horizon, df_returns, returns_neutral, returns_type)
    screen_full_feat, new_features = add_features_variation(screen_clean, features, variations_freq)

    screen_full_feat.to_pickle(df_features_path)

def save_mapping(file,strat_file_path,preprocessing_file_path):
    xl = pd.ExcelFile(file)
    strat_ML = xl.parse('Mapping_strat_ML', header=[0,1])
    preprocessing_df = xl.parse('Mapping_preprocessing', header=[0,1])

    unique_strat = list(strat_ML.columns.get_level_values(0).unique())
    dict_strat = {}
    for strat in unique_strat:
        dict_strat[strat] = {col: ((strat_ML[strat][col].dropna().values)[0] if len(list(strat_ML[strat][col].dropna().values))==1 else list(strat_ML[strat][col].dropna().values)) for col in strat_ML[strat].columns}

    unique_preprocessing = list(preprocessing_df.columns.get_level_values(0).unique())
    dict_preprocessing = {}
    for preprocessing in unique_preprocessing:
        dict_preprocessing[preprocessing] = {col: ((preprocessing_df[preprocessing][col].dropna().values)[0] if len(list(preprocessing_df[preprocessing][col].dropna().values))==1 else list(preprocessing_df[preprocessing][col].dropna().values)) for col in preprocessing_df[preprocessing].columns}

    file = open(strat_file_path, 'wb')
    pickle.dump(dict_strat, file)
    file.close()
    file = open(preprocessing_file_path, 'wb')
    pickle.dump(dict_preprocessing, file)
    file.close()

def read_pickle(file):

    """ Lit un fichier pickle """

    pkl_object = open(file, 'rb')
    object = pickle.load(pkl_object)
    pkl_object.close()
    return object