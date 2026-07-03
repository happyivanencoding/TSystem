import pandas as pd
import numpy as np
import xgboost as xgb
from math import *
from operator import itemgetter
import pickle
from dateutil import relativedelta

 
from sklearn.preprocessing import LabelEncoder, PolynomialFeatures
from sklearn.metrics import accuracy_score, f1_score
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import RandomizedSearchCV
from sklearn.impute import KNNImputer
from sklearn.pipeline import make_pipeline
from scipy.stats import rankdata
import shap

 
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import product
 
import copy

# import py_files.config as cf

import warnings
warnings.filterwarnings('ignore')

# def predict(df_scores, 
#             model_name, 
#             model_params, 
#             period_to_predict, 
#             include_columns, 
#             split_data_date, 
#             monotone_constraints, 
#             type_label = 'none', 
#             classes_weights=None):
#     """
#     Perform model training and prediction for multiple time periods.
    
#     Parameters:
#     -----------
#     df_scores : DataFrame
#         Input DataFrame with features and labels
#     model_name : str
#         Name of the model to use
#     model_params : dict
#         Model hyperparameters
#     period_to_predict : list
#         List of prediction periods (e.g., [3.0, 12.0] for 3-month and 12-month predictions)
#     include_columns : list
#         Features to use in the model
#     split_data_date : datetime
#         Date separating training and testing data
#     monotone_constraints : dict
#         Feature monotonicity constraints
#     sampling : str, default='base'
#         Sampling method for training data
#     obs_weight : str, default='none'
#         Observation weighting method
#     type_label : str, default='none'
#         Type of prediction task ('classification' or 'regression')
#     classes_weights : dict, optional
#         Weights for different classes in classification
#     """

#     # Split data into training and testing sets based on date
#     # Rolling trainning and testing
#     # if training_window_yr = 5, test_window_mth = 12, then model will use data of last 6 years, for the first 5 years as train set, for the last 12 months as test set
#     training_set = df_scores.loc[df_scores.index.get_level_values('Date') < split_data_date]
#     test_set = df_scores.loc[df_scores.index.get_level_values('Date') >= split_data_date]
   
#     y_pred=[] # This will contain predictions of returns in different months
#     for period in period_to_predict: # for prod, period_to_predict : [3.0, 12.0]
#         training_set_na_null = training_set.copy(deep=True).dropna(subset=[f"{period}M label"])
#         if type_label == 'classification':
#             le = LabelEncoder()
#             training_set[f"{period}M label"] = le.fit_transform(training_set[f"{period}M label"])
#             test_set[f"{period}M label"] = le.transform(test_set[f"{period}M label"])
 
#         X_train = training_set_na_null.drop(f"{period}M label", axis=1)
#         y_train = training_set_na_null[f"{period}M label"]

#         X_train = X_train[include_columns]
#         X_test = test_set[include_columns]
#         y_test = test_set[f"{period}M label"]
   
#         model = None # Initiation of final model
 
#         if model_name == 'XGBoostOptimizer':
#             if type_label == 'classification':
#                 xgb_model = xgb.XGBClassifier()
#             elif type_label == 'regression':
#                 xgb_model = xgb.XGBRegressor()
#             for param in model_params:
#                 if type(model_params[param])!=list:
#                     model_params[param] = [model_params[param]]
#             model = RandomizedSearchCV(estimator=xgb_model, param_distributions=model_params, n_iter=1000, cv=5, scoring='neg_mean_squared_error', random_state=42, n_jobs=-1)
#             #model.fit(X_train, y_train, sample_weight=classes_weights).best_params_
 
#         elif model_name == 'XGBoost' or model_name == 'XGBoost1' or model_name == 'XGBoost2':
#             if type_label=='classification':
#                 model = xgb.XGBClassifier(
#                 subsample=model_params['subsample'],
#                 reg_lambda=model_params['reg_lambda'],
#                 reg_alpha=model_params['reg_alpha'],
#                 random_state=42,
#                 n_estimators=model_params['n_estimators'],
#                 min_child_weight=model_params['min_child_weight'],
#                 max_depth=model_params['max_depth'],
#                 learning_rate=model_params['learning_rate'],
#                 gamma=model_params['gamma'],
#                 colsample_bytree=model_params['colsample_bytree'],
#                 monotone_constraints=monotone_constraints,
#                 objective = model_params['objective']
#                 #class_weight={0:4, 1:4, 2:1}
#             )
            
#             ############ Model à Utiliser #####################
#             elif type_label=='regression':
#                 model = xgb.XGBRegressor(
#                     subsample=model_params['subsample'],
#                     reg_lambda=model_params['reg_lambda'],
#                     reg_alpha=model_params['reg_alpha'],
#                     random_state=42,
#                     n_estimators=model_params['n_estimators'],
#                     min_child_weight=model_params['min_child_weight'],
#                     max_depth=model_params['max_depth'],
#                     learning_rate=model_params['learning_rate'],
#                     gamma=model_params['gamma'],
#                     colsample_bytree=model_params['colsample_bytree'],
#                     monotone_constraints=monotone_constraints,
#                     #class_weight={0:4, 1:4, 2:1}
#                     )
 
#         elif model_name == 'Logistic Regression Optimizer':
#             parameters = {
#                 'penalty': ['l2','none'],
#                 'C': [0.001, 0.01, 0.1, 1, 10, 100],
#                 'fit_intercept': [True, False],
#                 'max_iter': [50, 100, 200, 300, 500, 1000],
#                 'solver': ['newton-cg', 'lbfgs', 'sag', 'saga'],
#                 'random_state': [42],
#                 'class_weight': ['balanced']
#             }
 
#             lr_model = LogisticRegression()
#             model = RandomizedSearchCV(estimator=lr_model, param_distributions=parameters, n_iter=50, cv=3, scoring='f1_weighted', random_state=42)
#             print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train).best_params_)
 
#         elif model_name == 'Logistic Regression':
#             model = LogisticRegression(
#                 solver='liblinear',
#                 random_state=42,
#                 penalty='l2',
#                 max_iter=500,
#                 fit_intercept=True,
#                 class_weight='balanced',
#                 C=0.01
#             )
 
#         elif model_name == 'Random Forest Optimizer':
#             parameters = {
#                 'n_estimators': [int(x) for x in range(100, 1001, 100)],
#                 'criterion': ['gini', 'entropy', 'log_loss'],
#                 'max_depth': [int(x) for x in range(2, 6)],
#                 'min_samples_split': [2, 5, 10],
#                 'min_samples_leaf': [1, 2, 4],
#                 'bootstrap': [False, True],
#                 'warm_start': [False, True],
#                 'random_state' : [42]
#                 #'class_weight': ['balanced']
#             }
#             rf_model = RandomForestClassifier()
#             model = RandomizedSearchCV(estimator=rf_model, param_distributions=parameters, n_iter=50, cv=3, scoring='f1_weighted', random_state=42)
#             print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train, sample_weight=classes_weights).best_params_)
 
#         elif model_name == 'Random Forest':
#             model = RandomForestClassifier(
#                 warm_start=False,
#                 random_state=42,
#                 n_estimators=300,
#                 # min_samples_split=10,
#                 # min_samples_leaf=1,
#                 max_depth=5,
#                 criterion='gini',
#                 #class_weight='balanced',
#                 bootstrap=True
#             )
 
#         elif model_name == 'Multilinear Regression Optimizer':
#             poly = PolynomialFeatures(degree=2, include_bias=False)
#             linear_model = LinearRegression()
#             pipeline = make_pipeline(poly, linear_model)
 
#             parameters = {
#                 'polynomialfeatures__degree': [1, 2, 3],
#                 'linearregression__fit_intercept': [True, False],  
#             }
 
#             model = RandomizedSearchCV(
#                 pipeline,
#                 param_distributions=parameters,
#                 n_iter=10,
#                 scoring='neg_mean_squared_error',
#                 cv=5,
#                 random_state=42
#             )
#             print("Meilleurs hyperparamètres : ", model.fit(X_train, y_train).best_params_)
 
#         elif model_name == 'Multilinear Regression':
#             poly = PolynomialFeatures(degree=2, include_bias=False)
#             linear_model = LinearRegression(fit_intercept=True)
#             model = make_pipeline(poly, linear_model)
 
#         elif model_name == 'SVC':
#             model = SVC(
#                 class_weight='balanced',
#                 random_state=42
#             )

#         ### Model Fitting ###
#         if type_label == 'classification':
#             model.fit(X_train, y_train, sample_weight=classes_weights)
#         elif type_label == 'regression':
#             model.fit(X_train, y_train)

#         ### Model Prediction ###
#         if type_label == 'classification':
#             if 'Optimizer' in model_name:
#                 shap_explainer = shap.TreeExplainer(model.best_estimator_)
#             else:
#                 shap_explainer = shap.TreeExplainer(model)
#             shap_values = shap_explainer.shap_values(X_test)
#             class_names=list(le.inverse_transform(model.classes_))
#             shap_val_df = pd.DataFrame()
#             for i in range(len(class_names)):
#                 shap_val_df = pd.concat([shap_val_df, pd.DataFrame(data = np.abs(shap_values[i]).mean(0), index=include_columns, columns=[class_names[i]])], axis=1)
#             shap_val_df['overall'] = shap_val_df.sum(axis=1)
#             y_pred = model.predict(X_test)
#             y_proba = model.predict_proba(X_test)
#             f1 = f1_score(y_test, y_pred.round(), average='weighted')
#             accuracy = accuracy_score(y_test, y_pred.round())
#             if 'Regression' not in model_name:
#                 y_pred_labels = le.inverse_transform(y_pred)
#         elif type_label == 'regression':
#             if 'Optimizer' in model_name:
#                 shap_explainer = shap.TreeExplainer(model.best_estimator_)
#             else:
#                 shap_explainer = shap.TreeExplainer(model)
#             shap_values = shap_explainer.shap_values(X_test)
#             shap_val_df = pd.DataFrame(data = np.abs(shap_values).mean(0), index=include_columns, columns=['overall'])
#             new_pred = model.predict(X_test)
#             y_pred.append(new_pred)
#             # mse = mean_squared_error(y_test, new_pred)
#             # mse = 0
#             # r2 = r2_score(y_test, new_pred)
#             #r2_corr = 1 - (1- r2)*((len(y_test)-1)/(len(y_test)-len(include_columns)-1))
#             # r2_corr=0
 
#     if 'Optimizer' in model_name:
#         model_params = model.best_estimator_.get_params()
#     else:
#         model_params = model.get_params()
 
#     model_params = list(itemgetter('subsample','reg_lambda','reg_alpha','random_state','n_estimators','min_child_weight','max_depth','learning_rate','gamma','colsample_bytree','monotone_constraints')(model_params))
#     dict_params = dict(zip(['subsample','reg_lambda','reg_alpha','random_state','n_estimators','min_child_weight','max_depth','learning_rate','gamma','colsample_bytree','monotone_constraints'],model_params))
 
#     if type_label == 'classification':
#         return y_pred_labels,y_proba, shap_val_df, class_names, f1, accuracy, dict_params, [model,X_test]
#     elif type_label == 'regression':
#         return y_pred



# def dist_pred(y_true, y_pred):
 
#     ranks_true = rankdata(y_true)
#     ranks_true = len(ranks_true)-ranks_true+1
 
#     ranks_pred = rankdata(y_pred)
#     ranks_pred = len(ranks_pred)-ranks_pred+1
   
#     dist = np.sum(np.abs(ranks_true-ranks_pred))
 
#     return dist

# def export_results_to_excel(wb, sheet_name, perf, buy_list_all, parameters, score1, score2, shap_values, type_label):

#     wb.create_sheet(sheet_name)
#     ws = wb.get_sheet_by_name(sheet_name)
#     rows = dataframe_to_rows(perf, index=True)
#     for r_idx, row in enumerate(rows, 1):
#         for c_idx, value in enumerate(row, 1):
#             ws.cell(row=r_idx, column=c_idx, value=value)
#     if type_label == 'classification':
#         ws.cell(row=1,column=len(perf.columns)+3,value = 'f1 score')
#         ws.cell(row=2,column=len(perf.columns)+3,value = 'accuracy')
#     if type_label == 'regression':
#         ws.cell(row=1,column=len(perf.columns)+3,value = 'MSE')
#         ws.cell(row=2,column=len(perf.columns)+3,value = 'R2 corrigé')
#     ws.cell(row=1,column=len(perf.columns)+4,value = score1)
#     ws.cell(row=2,column=len(perf.columns)+4,value = score2)

#     rows = dataframe_to_rows(shap_values, index=True)
#     for r_idx, row in enumerate(rows, 1):
#         for c_idx, value in enumerate(row, 1):
#             ws.cell(row=3+r_idx, column=len(perf.columns)+2+c_idx, value=value)
    
#     ws.cell(row=7+len(shap_values),column=len(perf.columns)+3,value = 'parameters')
#     for i,param in enumerate(parameters):
#         ws.cell(row=7+len(shap_values)+i,column=len(perf.columns)+4,value = str(param))

#     width_buy_list = len(buy_list_all[list(buy_list_all)[0]].columns)+1
#     shift_col_buy_list = 0
#     for buy_list in buy_list_all.keys():
#         ws.cell(row=1, column=len(perf.columns)+len(shap_values.columns)+5+shift_col_buy_list, value=buy_list)
#         rows = dataframe_to_rows(buy_list_all[buy_list], index=False)
#         for r_idx, row in enumerate(rows, 1):
#             for c_idx, value in enumerate(row, 1):
#                 ws.cell(row=r_idx+1, column=c_idx +len(perf.columns)+len(shap_values.columns)+4+shift_col_buy_list, value=value)
#         shift_col_buy_list += width_buy_list

# def get_all_params_from_excel(Excel_Launcher_path):
#     params_principal = pd.read_excel(Excel_Launcher_path, header=0, index_col=0, sheet_name="Principal")
#     strat_ml_selected = params_principal.loc['strat ML', "param"]

#     params_preprocessing = pd.read_excel(Excel_Launcher_path, header=0, sheet_name="Preprocessing")

#     params_strat = pd.read_excel(Excel_Launcher_path, header=0, sheet_name="Strat_ML")
#     params_strat_selected = process_strategy_columns(params_strat, strat_ml_selected)

#     params_hyper_parameters = pd.read_excel(Excel_Launcher_path, header=0, index_col=0, sheet_name="Hyper_parameters")
#     ML_algo_selected = get_items_params_without_nan(params_strat_selected['algo'])
#     params_hyper_parameters_selected = params_hyper_parameters[ML_algo_selected]

#     return params_principal, params_preprocessing, params_strat_selected, params_hyper_parameters_selected

# def launch_prediction(input_transformed, params_strat_selected, params_hyper_parameters, classes=None):
#     # Unpacking values from strat_params
#     period_to_predict = get_items_params_without_nan(params_strat_selected['period_to_predict'])
#     returns_type = get_items_params_without_nan(params_strat_selected['Y'])
#     features = get_items_params_without_nan(params_strat_selected['X']) 
#     min_pct_avail_features = get_items_params_without_nan(params_strat_selected['min_avail_features'])
#     fill_na_method = get_items_params_without_nan(params_strat_selected['fill_na_X'])
#     model_name = get_items_params_without_nan(params_strat_selected['algo'])
#     sampling_method = get_items_params_without_nan(params_strat_selected['sampling_method'])
#     training_window = get_items_params_without_nan(params_strat_selected['training_window_yr'])
#     test_window = get_items_params_without_nan(params_strat_selected['test_window_mth'])
#     obs_weight = get_items_params_without_nan(params_strat_selected['obs_weight'])
#     type_label = get_items_params_without_nan(params_strat_selected['type_label'])

#     feature_constraints = dict(zip(get_items_params_without_nan(params_strat_selected['X']), get_items_params_without_nan(params_strat_selected['X_constraints'])))

#     screen_agg = copy.deepcopy(input_transformed)

#     screen_na_clean = fill_nan_values(screen_agg, features, period_to_predict, min_pct_avail_features, fill_na_method, obs_weight, returns_type)

#     screen_label = labellize_data(screen_na_clean,period_to_predict, returns_type, type_label)
#     screen_label = screen_label[screen_label['Weight in univ']>0]

#     # Most recent date
#     last_date = screen_label.index.get_level_values('Date').max()
#     # 5 years + 12 months ago (6 years)
#     cut_screen_date = last_date - relativedelta.relativedelta(years=training_window, months =test_window, day=1)
#     # Taking 6 years historical data for prediction
#     screen_label = screen_label[screen_label.index.get_level_values('Date') >= cut_screen_date]

#     # Find the earlist date for the past 6 years
#     first_date = screen_label.index.get_level_values('Date').min()
#     # Using the first 5 years as trainning data
#     split_date = first_date + relativedelta.relativedelta(years=training_window, months=1, day= 1)
#     # Taking the last year as testing data
#     test_dataset = screen_label.loc[screen_label.index.get_level_values('Date') >= split_date].reset_index()   

#     y_pred = predict(screen_label, model_name, params_hyper_parameters, period_to_predict, features, split_date, feature_constraints, sampling_method, obs_weight, type_label)

#     for i, period in enumerate(period_to_predict):
#                 test_dataset['predicted_return_'+str(period)+'M'] = y_pred[i]

#     col_predict=[]
#     for period in period_to_predict:
#         col_predict.append('predicted_return_'+str(period)+'M')
#         test_dataset['predicted_return_'+str(period)+'M'] = test_dataset['predicted_return_'+str(period)+'M'].rank(ascending=True)

#     test_dataset['Score ML'] = test_dataset[col_predict].mean(axis=1)
#     test_dataset['Score ML'] = test_dataset.groupby('Date')['Score ML'].rank(pct=True, ascending=True)*10

#     return test_dataset[test_dataset['Date'] == test_dataset['Date'].max()]
#     # return test_dataset

# def launch_ML_prediction(screen_agg_origin, period_to_predict, returns_type, features, min_pct_avail_features, fill_na_method, classes, model_name,model_params, sampling_method, feature_constraints, training_window, test_window, obs_weight, type_label):

#     screen_agg = copy.deepcopy(screen_agg_origin)

#     screen_na_clean = fill_nan_values(screen_agg, features, period_to_predict, min_pct_avail_features, fill_na_method, obs_weight, returns_type)
   
#     screen_label = labellize_data(screen_na_clean,period_to_predict,classes, returns_type, type_label)
#     screen_label = screen_label[screen_label['Weight in univ']>0]
 
#     last_date = screen_label.index.get_level_values('Date').max()
#     cut_screen_date = last_date - relativedelta.relativedelta(years=training_window, months =test_window, day=1)
#     screen_label = screen_label[screen_label.index.get_level_values('Date') >= cut_screen_date]

#     first_date = screen_label.index.get_level_values('Date').min()
#     split_date = first_date + relativedelta.relativedelta(years=training_window, months=1, day= 1)
#     test_dataset = screen_label.loc[screen_label.index.get_level_values('Date') >= split_date].reset_index()
   
#     y_pred, mse = predict(screen_label, model_name, model_params, period_to_predict, features, split_date, feature_constraints, sampling_method, obs_weight, type_label)

#     for i, period in enumerate(period_to_predict):
#                 test_dataset['predicted_return_'+str(period)+'M'] = y_pred[i]

#     col_predict=[]
#     for period in period_to_predict:
#         col_predict.append('predicted_return_'+str(period)+'M')
#         test_dataset['predicted_return_'+str(period)+'M'] = test_dataset['predicted_return_'+str(period)+'M'].rank(ascending=True)

#     test_dataset['Score ML'] = test_dataset[col_predict].mean(axis=1)
#     test_dataset['Score ML'] = test_dataset.groupby('Date')['Score ML'].rank(pct=True, ascending=True)*10

#     return test_dataset[test_dataset['Date'] == test_dataset['Date'].max()]

# def format_ML_params(df_params, algo):

#     param_algo = copy.deepcopy(df_params[algo])
#     for param_name,param_val in param_algo.items():
#         if ';' in str(param_val):
#             param_str = param_val.split(';')
#             param_algo[param_name] = [eval(p) for p in param_str]
#     return param_algo.to_dict()



# def save_mapping(file,strat_file_path,preprocessing_file_path):
#     xl = pd.ExcelFile(file)
#     strat_ML = xl.parse('Mapping_strat_ML', header=[0,1])
#     preprocessing_df = xl.parse('Mapping_preprocessing', header=[0,1])

#     unique_strat = list(strat_ML.columns.get_level_values(0).unique())
#     dict_strat = {}
#     for strat in unique_strat:
#         dict_strat[strat] = {col: ((strat_ML[strat][col].dropna().values)[0] if len(list(strat_ML[strat][col].dropna().values))==1 else list(strat_ML[strat][col].dropna().values)) for col in strat_ML[strat].columns}

#     unique_preprocessing = list(preprocessing_df.columns.get_level_values(0).unique())
#     dict_preprocessing = {}
#     for preprocessing in unique_preprocessing:
#         dict_preprocessing[preprocessing] = {col: ((preprocessing_df[preprocessing][col].dropna().values)[0] if len(list(preprocessing_df[preprocessing][col].dropna().values))==1 else list(preprocessing_df[preprocessing][col].dropna().values)) for col in preprocessing_df[preprocessing].columns}

#     file = open(strat_file_path, 'wb')
#     pickle.dump(dict_strat, file)
#     file.close()
#     file = open(preprocessing_file_path, 'wb')
#     pickle.dump(dict_preprocessing, file)
#     file.close()

# def read_pickle(file):

#     """ Lit un fichier pickle """

#     pkl_object = open(file, 'rb')
#     object = pickle.load(pkl_object)
#     pkl_object.close()
#     return object


